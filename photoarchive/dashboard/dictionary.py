"""What the dictionary currently holds, read from ``catalog.xlsx``.

The dashboard is a viewer, so it reads the same workbook a person opens rather
than reaching into SQLite. Two things are worth surfacing beside the photos:

* **candidates awaiting a decision** — machine guesses that do nothing until
  someone promotes or rejects them, and which are invisible unless you go
  looking;
* **what the last learn invented** — new canonical entities. A typed name that
  matches nothing becomes a new person, place or tag, so a typo turns into a
  dictionary entry. Noticing that on the same page as the photos is the point;
  by the time it is noticed in `catalog.xlsx`, it has usually propagated.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

CATALOG_FILENAME = "catalog.xlsx"

#: ``sheet -> (canonical column, kind label)``.
_ENTITY_SHEETS: tuple[tuple[str, str, str], ...] = (
    ("People", "canonical_name", "people"),
    ("Places", "canonical_place", "places"),
    ("Tags", "canonical_tag", "tags"),
)


@dataclass(slots=True)
class DictionarySummary:
    """Counts and warnings drawn from the catalog workbook."""

    people: int = 0
    places: int = 0
    tags: int = 0
    candidates: int = 0
    rejected: int = 0
    #: Canonical values whose earliest evidence comes from the most recent run.
    created_recently: list[str] = field(default_factory=list)
    #: True when no catalog workbook was found beside the review workbooks.
    missing: bool = True

    @property
    def entities(self) -> int:
        return self.people + self.places + self.tags

    @property
    def needs_attention(self) -> bool:
        return bool(self.candidates or self.created_recently)


def read_summary(source_dir: Path | str) -> DictionarySummary:
    """Summarise ``catalog.xlsx``; a missing or unreadable one summarises to nothing."""
    path = Path(source_dir) / CATALOG_FILENAME
    if not path.exists():
        return DictionarySummary()

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - the dashboard must render regardless
        return DictionarySummary()

    try:
        summary = DictionarySummary(missing=False)
        for sheet_name, canonical_column, kind in _ENTITY_SHEETS:
            if sheet_name not in workbook.sheetnames:
                continue
            rows = _rows(workbook[sheet_name])
            setattr(summary, kind, len(rows))
            for row in rows:
                summary.candidates += len(_values(row.get("candidate_aliases")))
                summary.rejected += len(_values(row.get("rejected_aliases")))
            del canonical_column  # named for readability of the mapping above

        summary.created_recently = _created_in_latest_run(workbook)
        return summary
    finally:
        workbook.close()


def _created_in_latest_run(workbook) -> list[str]:
    """Entities whose provenance begins in the newest run present.

    An entity created earlier has older evidence too, so comparing each
    entity's *first* run against the newest one distinguishes "invented just
    now" from "mentioned again".
    """
    if "Evidence" not in workbook.sheetnames:
        return []

    rows = _rows(workbook["Evidence"])
    runs = [row.get("run_id", "") for row in rows if row.get("run_id")]
    if not runs:
        return []

    latest = max(runs)
    first_run: dict[str, str] = {}
    for row in rows:
        value, run = row.get("entity_value", ""), row.get("run_id", "")
        if not value or not run:
            continue
        if value not in first_run or run < first_run[value]:
            first_run[value] = run

    return sorted(value for value, run in first_run.items() if run == latest)


def _rows(sheet) -> list[dict[str, str]]:
    raw = list(sheet.iter_rows(values_only=True))
    if not raw:
        return []
    headers = [str(name) if name else "" for name in raw[0]]
    result: list[dict[str, str]] = []
    for values in raw[1:]:
        row = {
            headers[index]: ("" if value is None else str(value).strip())
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        if any(row.values()):
            result.append(row)
    return result


def _values(cell: str | None) -> list[str]:
    return [part.strip() for part in (cell or "").split(";") if part.strip()]


__all__ = ["CATALOG_FILENAME", "DictionarySummary", "read_summary"]
