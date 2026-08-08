"""``review.xlsx``: schema, reading, and generation with embedded previews.

Invariants of this workbook:

1. **One workbook per folder that directly contains photos.**
2. **Previews are embedded**, so a reviewer never opens the images separately.
   ``DESCRIBED_ABSENT`` rows exist with an empty Preview cell — described, but
   no photo to show.
3. **Suggested columns are machine-owned; the final columns are not.** A rescan
   may rewrite ``Suggested …`` freely and must never write ``Date``, ``Place``,
   ``LatLon``, ``People`` or ``Tags`` on a row that already existed.
4. **Source changes are surfaced, not applied** — the row returns to review
   with a stated ``Review Reason``, keeping whatever the reviewer had entered.
5. **Row order is stable** across unchanged rescans, so two runs diff cleanly.

Coordinates are text in the canonical ``lat, lon`` form, and both coordinate
cells carry a Google Maps hyperlink on the cell itself rather than in an extra
visible column.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from photoarchive.geo import parse_latlon
from photoarchive.models import WorkflowStatus
from photoarchive.naming import filename_stem
from photoarchive.review.model import ReviewRow

LOG = logging.getLogger(__name__)

SHEET_NAME = "Review"

REVIEW_FILENAME = "review.xlsx"

#: On-sheet column order. Each machine-owned "Suggested X" sits immediately
#: before the user-owned "X" it seeds, so the pair reads as one decision.
VISIBLE_COLUMNS: tuple[str, ...] = (
    "Preview",
    "Filename / Reference",
    "Source Description",
    "Section Context",
    "Source Notes",
    "Suggested Date",
    "Date",
    "Suggested Place",
    "Place",
    "Suggested LatLon",
    "LatLon",
    "Map Link",
    "Suggested People",
    "People",
    "Suggested Tags",
    "Tags",
    "Event",
    "Albums",
    "Description",
    "Status",
    "Review Reason",
    "Notes",
)

#: Machine-owned: safe to rewrite on any scan.
SUGGESTED_COLUMNS: frozenset[str] = frozenset(
    {
        "Suggested Date",
        "Suggested Place",
        "Suggested LatLon",
        "Suggested People",
        "Suggested Tags",
    }
)

#: User-owned: written once at creation, never overwritten afterwards.
FINAL_COLUMNS: frozenset[str] = frozenset(
    {"Date", "Place", "LatLon", "People", "Tags"}
)

#: Attribute on :class:`ReviewRow` backing each column.
_COLUMN_ATTRIBUTES: dict[str, str] = {
    "Filename / Reference": "filename_or_reference",
    "Source Description": "source_description",
    "Section Context": "section_context",
    "Source Notes": "source_notes",
    "Suggested Date": "suggested_date",
    "Date": "date",
    "Suggested Place": "suggested_place",
    "Place": "place",
    "Suggested LatLon": "suggested_latlon",
    "LatLon": "latlon",
    "Map Link": "map_link",
    "Suggested People": "suggested_people",
    "People": "people",
    "Suggested Tags": "suggested_tags",
    "Tags": "tags",
    "Event": "event",
    "Albums": "albums",
    "Description": "description",
    "Review Reason": "review_reason",
    "Notes": "notes",
}

_COLUMN_WIDTHS: dict[str, int] = {
    "Preview": 26,
    "Filename / Reference": 22,
    "Source Description": 60,
    "Section Context": 24,
    "Source Notes": 16,
    "Map Link": 30,
    "Description": 40,
    "Review Reason": 30,
    "Notes": 24,
}
_DEFAULT_WIDTH = 18

_WRAPPED_COLUMNS = frozenset(
    {"Source Description", "Section Context", "Description", "Review Reason", "Notes"}
)

PREVIEW_WIDTH_PX = 180
#: Excel row heights are in points; 1 px is 0.75 pt.
_PX_TO_POINTS = 0.75

_HEADER_FILL = PatternFill("solid", fgColor="DDE5F0")
_SUGGESTED_FILL = PatternFill("solid", fgColor="F2F2F2")
_HYPERLINK_FONT = Font(color="0563C1", underline="single")


def column_index(name: str) -> int:
    """Return the 1-based worksheet column index for a schema column."""
    return VISIBLE_COLUMNS.index(name) + 1


def identity_key(value: str) -> str:
    """Row identity *within one folder*: the reference without its extension.

    ``20200512_150442``, ``20200512_150442.jpg`` and ``IMG_001.jpeg`` all
    reduce to one key, so a described-but-absent row and the photo that later
    turns up are recognised as the same row.

    This key is deliberately folder-local. Two folders may each hold a
    ``001.jpg``, and they are different photos — see :func:`scoped_row_key` for
    the archive-wide identity.
    """
    return filename_stem(str(value).strip()).casefold()


def scoped_row_key(root_identity: str, folder_path: str, reference: str) -> str:
    """Archive-wide row identity: source root + folder + folder-local key.

    A bare stem is not an identity across the archive: ``folder A/001.jpg`` and
    ``folder B/001.jpg`` are unrelated photos. Anything that stores rows beyond
    a single workbook — the SQLite bookkeeping, and any future cross-folder
    search — must key on this, not on the stem alone.
    """
    folder = "/".join(part for part in str(folder_path).split("/") if part)
    return f"{root_identity}|{folder}|{identity_key(reference)}"


@dataclass(frozen=True, slots=True)
class WorkbookPreview:
    """A thumbnail to embed, already sized and written to a local file."""

    reference: str
    image_path: Path
    width_px: int
    height_px: int


class ReviewWorkbookService:
    """Reads and writes one folder's ``review.xlsx``.

    Operates purely on local files; uploading is a later phase's problem.
    """

    def __init__(self, preview_width_px: int = PREVIEW_WIDTH_PX) -> None:
        self.preview_width_px = preview_width_px

    # -- Reading ----------------------------------------------------------

    def read(self, path: Path) -> dict[str, ReviewRow]:
        """Load existing rows, keyed by :func:`identity_key`.

        Identity is the reference stem rather than the printed cell, so a row
        keeps its identity when a ``DESCRIBED_ABSENT`` reference later gains a
        filename. That is what lets such a row be reused instead of duplicated,
        without spending a visible bookkeeping column on it.

        Returns an empty mapping when the workbook does not exist yet. Unknown
        or reordered columns are tolerated: lookup is by header name.
        """
        path = Path(path)
        if not path.exists():
            return {}

        workbook = load_workbook(path)
        worksheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {str(name): index for index, name in enumerate(header_row) if name}

        rows: dict[str, ReviewRow] = {}
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            reference = _cell(values, headers, "Filename / Reference")
            if not reference:
                continue

            row = ReviewRow(reference=reference)
            for column, attribute in _COLUMN_ATTRIBUTES.items():
                if attribute == "filename_or_reference":
                    continue
                setattr(row, attribute, _cell(values, headers, column))

            status = _cell(values, headers, "Status")
            row.status = _parse_status(status)
            rows[identity_key(reference)] = row

        workbook.close()
        return rows

    # -- Writing ----------------------------------------------------------

    def write(
        self,
        path: Path,
        rows: list[ReviewRow],
        previews: dict[str, WorkbookPreview] | None = None,
    ) -> Path:
        """Write the workbook, embedding previews for rows that have one."""
        from openpyxl.drawing.image import Image as ExcelImage

        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        previews = previews or {}

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME

        self._write_header(worksheet)

        for offset, row in enumerate(rows):
            excel_row = offset + 2
            self._write_row(worksheet, excel_row, row)

            preview = previews.get(row.reference) or previews.get(
                identity_key(row.reference)
            )
            if preview is None or not preview.image_path.exists():
                # DESCRIBED_ABSENT and unreadable photos simply have no image.
                worksheet.row_dimensions[excel_row].height = 18
                continue

            image = ExcelImage(str(preview.image_path))
            image.width, image.height = preview.width_px, preview.height_px
            worksheet.add_image(image, f"A{excel_row}")
            worksheet.row_dimensions[excel_row].height = (
                preview.height_px * _PX_TO_POINTS + 6
            )

        self._apply_layout(worksheet, len(rows))
        workbook.save(path)
        workbook.close()
        return path

    def _write_header(self, worksheet) -> None:
        for index, name in enumerate(VISIBLE_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=index, value=name)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def _write_row(self, worksheet, excel_row: int, row: ReviewRow) -> None:
        for column in VISIBLE_COLUMNS:
            index = column_index(column)
            if column == "Preview":
                continue

            if column == "Filename / Reference":
                value = row.filename or row.reference
            elif column == "Status":
                value = row.status.value
            else:
                value = getattr(row, _COLUMN_ATTRIBUTES[column], "") or ""

            cell = worksheet.cell(row=excel_row, column=index, value=value)
            if column in _WRAPPED_COLUMNS:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")
            if column in SUGGESTED_COLUMNS:
                cell.fill = _SUGGESTED_FILL

            # Both coordinate cells link to the point; the text stays canonical.
            if column in ("Suggested LatLon", "LatLon") and value:
                point = parse_latlon(value)
                if point is not None:
                    cell.hyperlink = point.map_url
                    cell.font = _HYPERLINK_FONT

    def _apply_layout(self, worksheet, row_count: int) -> None:
        for index, name in enumerate(VISIBLE_COLUMNS, start=1):
            letter = get_column_letter(index)
            worksheet.column_dimensions[letter].width = _COLUMN_WIDTHS.get(
                name, _DEFAULT_WIDTH
            )

        worksheet.freeze_panes = "C2"
        last_column = get_column_letter(len(VISIBLE_COLUMNS))
        worksheet.auto_filter.ref = f"A1:{last_column}{max(row_count + 1, 1)}"


def build_preview(
    source_image: Path, destination: Path, width_px: int = PREVIEW_WIDTH_PX
) -> tuple[int, int] | None:
    """Downscale one photo into a thumbnail file, leaving the source untouched.

    Returns the thumbnail's pixel size, or ``None`` when the image could not be
    read — an unreadable photo costs a preview, never the whole workbook.
    """
    from PIL import Image, UnidentifiedImageError

    try:
        with Image.open(source_image) as picture:
            picture = picture.convert("RGB")
            ratio = width_px / float(picture.width)
            size = (width_px, max(1, int(picture.height * ratio)))
            thumbnail = picture.resize(size)
            destination.parent.mkdir(parents=True, exist_ok=True)
            thumbnail.save(destination, format="PNG")
            return size
    except (UnidentifiedImageError, OSError, ValueError) as error:
        LOG.debug("Could not build a preview for %s: %s", source_image.name, error)
        return None


def _cell(values: tuple, headers: dict[str, int], column: str) -> str:
    index = headers.get(column)
    if index is None or index >= len(values):
        return ""
    value = values[index]
    return "" if value is None else str(value).strip()


def _parse_status(value: str) -> WorkflowStatus:
    try:
        return WorkflowStatus(value)
    except ValueError:
        return WorkflowStatus.NEW
