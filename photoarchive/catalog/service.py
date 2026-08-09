"""The global ``catalog.xlsx``: the dictionaries in human-readable form.

SQLite is where the dictionaries live; this workbook is the view a person
opens. It carries three sheets — ``People``, ``Places`` and ``Tags`` — and its
job is to make one distinction impossible to miss:

* **CONFIRMED** knowledge the pipeline may act on;
* **CANDIDATE** knowledge that is only a hint, shaded amber and never used for
  a suggestion until a human promotes it.

Each row also shows how much evidence stands behind it, so "why does the system
think this?" is answerable without opening the database.

In this milestone the workbook is written, not read back: SQLite remains the
source of truth, and promoting a candidate is a deliberate action rather than
an Excel edit.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from photoarchive.catalog.models import ConfidenceStatus, Dictionary, EntityType
from photoarchive.catalog.store import DictionaryStore

CATALOG_FILENAME = "catalog.xlsx"

CATALOG_SHEETS: tuple[str, ...] = ("People", "Places", "Tags")

PEOPLE_COLUMNS: tuple[str, ...] = (
    "person_id",
    "canonical_name",
    "confirmed_aliases",
    "candidate_aliases",
    "evidence_count",
    "notes",
)

PLACES_COLUMNS: tuple[str, ...] = (
    "place_id",
    "canonical_place",
    "confirmed_aliases",
    "candidate_aliases",
    "latlon",
    "candidate_latlon",
    "map_link",
    "evidence_count",
    "notes",
)

TAGS_COLUMNS: tuple[str, ...] = (
    "tag_id",
    "canonical_tag",
    "confirmed_aliases",
    "candidate_aliases",
    "evidence_count",
    "notes",
)

_HEADER_FILL = PatternFill("solid", fgColor="DDE5F0")
#: Amber: this cell holds a hint, not a fact.
_CANDIDATE_FILL = PatternFill("solid", fgColor="FFF2CC")

_LIST_SEPARATOR = "; "

_CANDIDATE_COLUMNS = frozenset({"candidate_aliases", "candidate_latlon"})


@dataclass(frozen=True, slots=True)
class CatalogCounts:
    """Summary of what was exported."""

    people: int = 0
    places: int = 0
    tags: int = 0
    candidate_aliases: int = 0
    candidate_coordinates: int = 0


class CatalogService:
    """Exports the dictionaries to ``catalog.xlsx``."""

    def __init__(self, filename: str = CATALOG_FILENAME) -> None:
        self.filename = filename
        #: False when the last export left an already-correct file untouched.
        self.written = True

    def export(
        self, store: DictionaryStore, output_dir: Path
    ) -> tuple[Path, CatalogCounts]:
        """Write ``catalog.xlsx``, unless its content is already on disk.

        Like ``review.xlsx``, this is a ZIP container: rewriting it changes the
        bytes even when the dictionary did not, which would mean a pointless
        upload and a pointless portable-state generation on every run.
        """
        dictionary = store.load()
        path = Path(output_dir) / self.filename
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)

        counts = CatalogCounts(
            people=len(dictionary.people),
            places=len(dictionary.places),
            tags=len(dictionary.tags),
            candidate_aliases=_count_candidate_aliases(dictionary),
            candidate_coordinates=sum(
                len(place.candidate_latlon) for place in dictionary.places
            ),
        )

        self._write_people(workbook, store, dictionary)
        self._write_places(workbook, store, dictionary)
        self._write_tags(workbook, store, dictionary)

        if _signature(workbook) == _existing_signature(path):
            workbook.close()
            self.written = False
            return path, counts

        workbook.save(path)
        workbook.close()
        self.written = True
        return path, counts

    def _write_people(self, workbook, store, dictionary: Dictionary) -> None:
        sheet = workbook.create_sheet("People")
        _write_header(sheet, PEOPLE_COLUMNS)
        for index, person in enumerate(dictionary.people, start=2):
            _write_row(
                sheet,
                index,
                PEOPLE_COLUMNS,
                {
                    "person_id": person.person_id,
                    "canonical_name": person.canonical_name,
                    "confirmed_aliases": _join(person.confirmed_aliases),
                    "candidate_aliases": _join(person.candidate_aliases),
                    "evidence_count": store.evidence_count(
                        EntityType.PERSON, person.canonical_name
                    ),
                    "notes": person.notes or "",
                },
            )
        _size(sheet, PEOPLE_COLUMNS)

    def _write_places(self, workbook, store, dictionary: Dictionary) -> None:
        sheet = workbook.create_sheet("Places")
        _write_header(sheet, PLACES_COLUMNS)
        for index, place in enumerate(dictionary.places, start=2):
            _write_row(
                sheet,
                index,
                PLACES_COLUMNS,
                {
                    "place_id": place.place_id,
                    "canonical_place": place.canonical_place,
                    "confirmed_aliases": _join(place.confirmed_aliases),
                    "candidate_aliases": _join(place.candidate_aliases),
                    "latlon": place.latlon.format() if place.latlon else "",
                    "candidate_latlon": _join(
                        point.format() for point in place.candidate_latlon
                    ),
                    "map_link": place.map_link or "",
                    "evidence_count": store.evidence_count(
                        EntityType.PLACE, place.canonical_place
                    ),
                    "notes": place.notes or "",
                },
            )
        _size(sheet, PLACES_COLUMNS)

    def _write_tags(self, workbook, store, dictionary: Dictionary) -> None:
        sheet = workbook.create_sheet("Tags")
        _write_header(sheet, TAGS_COLUMNS)
        for index, tag in enumerate(dictionary.tags, start=2):
            _write_row(
                sheet,
                index,
                TAGS_COLUMNS,
                {
                    "tag_id": tag.tag_id,
                    "canonical_tag": tag.canonical_tag,
                    "confirmed_aliases": _join(tag.confirmed_aliases),
                    "candidate_aliases": _join(tag.candidate_aliases),
                    "evidence_count": store.evidence_count(
                        EntityType.TAG, tag.canonical_tag
                    ),
                    "notes": tag.notes or "",
                },
            )
        _size(sheet, TAGS_COLUMNS)


def _signature(workbook) -> str:
    """Hash every cell of every sheet: the workbook's logical content."""
    digest = hashlib.sha256()
    for name in workbook.sheetnames:
        digest.update(name.encode("utf-8"))
        for row in workbook[name].iter_rows(values_only=True):
            for value in row:
                digest.update(str("" if value is None else value).encode("utf-8"))
                digest.update(b"\x1f")
            digest.update(b"\x1e")
    return digest.hexdigest()


def _existing_signature(path: Path) -> str | None:
    """The signature of the workbook already on disk, if it is readable."""
    if not Path(path).exists():
        return None
    try:
        workbook = load_workbook(path, read_only=True)
    except Exception:  # noqa: BLE001 - an unreadable file is simply rewritten
        return None
    try:
        return _signature(workbook)
    finally:
        workbook.close()


def _join(values) -> str:
    """Render a list field as a semicolon-separated cell value."""
    return _LIST_SEPARATOR.join(str(value) for value in values if value)


def _count_candidate_aliases(dictionary: Dictionary) -> int:
    total = sum(len(person.candidate_aliases) for person in dictionary.people)
    total += sum(len(place.candidate_aliases) for place in dictionary.places)
    total += sum(len(tag.candidate_aliases) for tag in dictionary.tags)
    return total


def _write_header(sheet, columns: tuple[str, ...]) -> None:
    for index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    sheet.freeze_panes = "A2"


def _write_row(sheet, row_index: int, columns: tuple[str, ...], values: dict) -> None:
    for index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=row_index, column=index, value=values.get(name, ""))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        # Shade candidate cells that actually hold something, so a reviewer
        # can never mistake a hint for confirmed knowledge.
        if name in _CANDIDATE_COLUMNS and values.get(name):
            cell.fill = _CANDIDATE_FILL


def _size(sheet, columns: tuple[str, ...]) -> None:
    for index, name in enumerate(columns, start=1):
        width = 40 if "alias" in name or "latlon" in name else 22
        sheet.column_dimensions[get_column_letter(index)].width = width


def status_label(status: ConfidenceStatus) -> str:
    """Human label for a confidence status."""
    return status.value
