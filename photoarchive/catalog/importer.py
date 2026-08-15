"""Importing human edits from ``catalog.xlsx`` back into SQLite.

The catalog is a two-way document: the pipeline writes it, a person edits it,
and the next run reads those edits back. That makes curation possible in the
tool people already have open — moving an alias from ``candidate_aliases`` to
``confirmed_aliases`` promotes it, typing coordinates into ``latlon`` confirms
them.

Three rules keep the round trip safe:

* **Stable ids win over names.** A row is matched by ``person_id`` /
  ``place_id`` / ``tag_id``, so renaming a canonical value edits that entity
  instead of creating a second one.
* **Human edits beat generated values** — but only after validating them.
* **Invalid input never destroys good data.** A malformed coordinate is
  reported and skipped; the previously confirmed value stays exactly as it was.

Evidence is never rewritten here. Promoting a candidate keeps every reason
already recorded for it and adds one for the promotion.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from photoarchive.catalog.models import (
    ConfidenceStatus,
    EntityType,
    Evidence,
    EvidenceReason,
)
from photoarchive.catalog.store import DictionaryStore
from photoarchive.geo import parse_latlon
from photoarchive.review.model import split_values

LOG = logging.getLogger(__name__)


@dataclass(slots=True)
class ImportOutcome:
    """What importing catalog edits changed, and what it refused."""

    entities_renamed: list[str] = field(default_factory=list)
    entities_added: list[str] = field(default_factory=list)
    aliases_confirmed: list[str] = field(default_factory=list)
    aliases_added_as_candidates: list[str] = field(default_factory=list)
    #: Spellings a person ruled out; matching skips them and learning never
    #: revives them.
    aliases_rejected: list[str] = field(default_factory=list)
    #: "<alias> -> <entity>" for an alias listed in two columns at once. A slip
    #: to report, not a state: rejection wins, and the person is told.
    collisions: list[str] = field(default_factory=list)
    promotions: list[str] = field(default_factory=list)
    #: "<absorbed> -> <survivor>" for duplicate entities merged into one.
    merged: list[str] = field(default_factory=list)
    latlon_updated: list[str] = field(default_factory=list)
    latlon_promoted: list[str] = field(default_factory=list)
    invalid: list[str] = field(default_factory=list)

    @property
    def changed(self) -> int:
        return (
            len(self.entities_renamed)
            + len(self.entities_added)
            + len(self.aliases_confirmed)
            + len(self.aliases_rejected)
            + len(self.promotions)
            + len(self.merged)
            + len(self.latlon_updated)
            + len(self.latlon_promoted)
        )


def import_catalog(
    store: DictionaryStore, path: Path, run_id: str | None = None
) -> ImportOutcome:
    """Read ``catalog.xlsx`` and sync validated human edits into SQLite.

    A missing workbook is not an error: there is simply nothing to import yet.
    """
    outcome = ImportOutcome()
    path = Path(path)
    if not path.exists():
        LOG.debug("No catalog workbook to import at %s", path)
        return outcome

    workbook = load_workbook(path, read_only=True)
    try:
        if "People" in workbook.sheetnames:
            _import_people(store, workbook["People"], outcome, run_id)
        if "Places" in workbook.sheetnames:
            _import_places(store, workbook["Places"], outcome, run_id)
        if "Tags" in workbook.sheetnames:
            _import_tags(store, workbook["Tags"], outcome, run_id)
    finally:
        workbook.close()

    return outcome


def _rows(sheet) -> list[dict[str, str]]:
    raw = list(sheet.iter_rows(values_only=True))
    if not raw:
        return []
    headers = [str(name) if name else "" for name in raw[0]]
    result: list[dict[str, str]] = []
    for values in raw[1:]:
        row = {
            headers[index]: ("" if value is None else str(value).strip())
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        if any(row.values()):
            result.append(row)
    return result


def _import_people(store, sheet, outcome: ImportOutcome, run_id) -> None:
    dictionary = store.load()
    by_id = {person.person_id: person for person in dictionary.people}

    for row in _rows(sheet):
        canonical = row.get("canonical_name", "")
        if not canonical:
            outcome.invalid.append("People: row without a canonical_name")
            continue

        entity_id = _resolve(
            store, EntityType.PERSON, row.get("person_id", ""), canonical, by_id,
            outcome, store.add_person,
        )
        _sync_aliases(
            store, EntityType.PERSON, entity_id, canonical, row,
            by_id.get(row.get("person_id", "")), outcome, run_id,
        )


def _import_tags(store, sheet, outcome: ImportOutcome, run_id) -> None:
    dictionary = store.load()
    by_id = {tag.tag_id: tag for tag in dictionary.tags}

    for row in _rows(sheet):
        canonical = row.get("canonical_tag", "")
        if not canonical:
            outcome.invalid.append("Tags: row without a canonical_tag")
            continue

        entity_id = _resolve(
            store, EntityType.TAG, row.get("tag_id", ""), canonical, by_id,
            outcome, store.add_tag,
        )
        _sync_aliases(
            store, EntityType.TAG, entity_id, canonical, row,
            by_id.get(row.get("tag_id", "")), outcome, run_id,
        )


def _import_places(store, sheet, outcome: ImportOutcome, run_id) -> None:
    dictionary = store.load()
    by_id = {place.place_id: place for place in dictionary.places}

    for row in _rows(sheet):
        canonical = row.get("canonical_place", "")
        if not canonical:
            outcome.invalid.append("Places: row without a canonical_place")
            continue

        place_id = _resolve(
            store, EntityType.PLACE, row.get("place_id", ""), canonical, by_id,
            outcome, store.add_place,
        )
        existing = by_id.get(row.get("place_id", ""))
        _sync_aliases(
            store, EntityType.PLACE, place_id, canonical, row, existing, outcome, run_id
        )
        _sync_place_coordinates(store, place_id, canonical, row, existing, outcome, run_id)


def _resolve(store, entity_type, entity_id, canonical, by_id, outcome, factory) -> str:
    """Find the entity this row refers to, honouring a rename.

    The id is authoritative: if it exists and the canonical value differs, the
    person renamed the entity and we update it in place rather than creating a
    duplicate under the new name.
    """
    existing = by_id.get(entity_id) if entity_id else None
    if existing is None:
        new_id = factory(canonical)
        if entity_id and entity_id != new_id:
            outcome.invalid.append(
                f"{entity_type.value}: unknown id {entity_id!r}; added {canonical!r} instead"
            )
        else:
            outcome.entities_added.append(canonical)
        return new_id

    current = _canonical_of(existing, entity_type)
    if current != canonical:
        store.rename_entity(entity_type, entity_id, canonical)
        outcome.entities_renamed.append(f"{current} -> {canonical}")
    return entity_id


def _canonical_of(entity, entity_type: EntityType) -> str:
    if entity_type is EntityType.PERSON:
        return entity.canonical_name
    if entity_type is EntityType.PLACE:
        return entity.canonical_place
    return entity.canonical_tag


def _sync_aliases(
    store, entity_type, entity_id, canonical, row, existing, outcome, run_id
) -> None:
    """Apply the alias columns as the person left them.

    The column an alias sits in *is* the decision, so all three are read as
    positive statements. An emptied cell says nothing: `catalog.xlsx` is
    three-way merged, and another machine's copy may simply not have seen a
    value yet — reading "gone" as "rejected" would turn a sync artefact into a
    decision.
    """
    confirmed = split_values(row.get("confirmed_aliases", ""))
    candidates = split_values(row.get("candidate_aliases", ""))
    rejected = split_values(row.get("rejected_aliases", ""))

    previously_candidate = set(existing.candidate_aliases) if existing else set()
    previously_confirmed = set(existing.confirmed_aliases) if existing else set()
    previously_rejected = set(existing.rejected_aliases) if existing else set()

    # One alias in two columns is a slip, not a state. Rejection wins: acting on
    # a spelling someone forbade is worse than ignoring one they allowed, and
    # the collision is reported rather than silently resolved.
    rejected_set = set(rejected)
    for alias in sorted(rejected_set.intersection(confirmed).union(
        rejected_set.intersection(candidates)
    )):
        outcome.collisions.append(f"{alias} -> {canonical}")
    confirmed = [alias for alias in confirmed if alias not in rejected_set]
    candidates = [alias for alias in candidates if alias not in rejected_set]

    for alias in rejected:
        if alias in previously_rejected:
            continue
        store.set_alias_status(
            entity_type, entity_id, alias, ConfidenceStatus.REJECTED
        )
        outcome.aliases_rejected.append(f"{alias} -> {canonical}")
        store.record_evidence(
            Evidence(
                entity_type=entity_type,
                entity_value=canonical,
                reason=EvidenceReason.MANUAL_CORRECTION,
                candidate_text=alias,
                status=ConfidenceStatus.REJECTED,
                run_id=run_id,
                source_folder="catalog.xlsx",
            )
        )

    for alias in confirmed:
        if alias in previously_confirmed:
            continue

        # Listing another entity's canonical value as an alias here is how a
        # person says "these two rows are the same thing". Merge rather than
        # keeping a duplicate that would reappear on the next export.
        duplicate_id = store.find_by_canonical(entity_type, alias)
        if duplicate_id and duplicate_id != entity_id:
            absorbed = store.merge_entities(entity_type, entity_id, duplicate_id)
            if absorbed:
                outcome.merged.append(f"{absorbed} -> {canonical}")
                continue

        if alias in previously_rejected:
            # Moved back out of the rejected column: the person changed their
            # mind, and `add_alias` would refuse to undo a rejection.
            store.set_alias_status(
                entity_type, entity_id, alias, ConfidenceStatus.CONFIRMED
            )
        else:
            store.add_alias(entity_type, entity_id, alias, ConfidenceStatus.CONFIRMED)
        if alias in previously_candidate:
            # Moved out of the candidate column: an explicit promotion.
            outcome.promotions.append(f"{alias} -> {canonical}")
        else:
            outcome.aliases_confirmed.append(f"{alias} -> {canonical}")
        store.record_evidence(
            Evidence(
                entity_type=entity_type,
                entity_value=canonical,
                reason=EvidenceReason.MANUAL_CORRECTION,
                candidate_text=alias,
                status=ConfidenceStatus.CONFIRMED,
                run_id=run_id,
                source_folder="catalog.xlsx",
            )
        )

    for alias in candidates:
        if alias in previously_candidate or alias in previously_confirmed:
            continue
        if alias in previously_rejected:
            store.set_alias_status(
                entity_type, entity_id, alias, ConfidenceStatus.CANDIDATE
            )
        else:
            store.add_alias(entity_type, entity_id, alias, ConfidenceStatus.CANDIDATE)
        outcome.aliases_added_as_candidates.append(f"{alias} -> {canonical}")


def _sync_place_coordinates(
    store, place_id, canonical, row, existing, outcome, run_id
) -> None:
    """Apply edits to a place's confirmed coordinates.

    An unparsable value is reported and ignored; the confirmed coordinate that
    was already there survives untouched.
    """
    raw = row.get("latlon", "")
    previous = existing.latlon if existing else None

    if not raw:
        return

    point = parse_latlon(raw)
    if point is None:
        outcome.invalid.append(f"Places: {canonical!r} has an unparsable latlon {raw!r}")
        return

    if previous is not None and previous.format() == point.format():
        return

    was_candidate = bool(
        existing and any(item.format() == point.format() for item in existing.candidate_latlon)
    )
    store.set_place_latlon(place_id, point)
    store.clear_place_latlon_candidate(place_id, point)

    if was_candidate:
        outcome.latlon_promoted.append(f"{canonical}: {point.format()}")
    else:
        outcome.latlon_updated.append(f"{canonical}: {point.format()}")

    store.record_evidence(
        Evidence(
            entity_type=EntityType.PLACE,
            entity_value=canonical,
            reason=EvidenceReason.MANUAL_CORRECTION,
            proposed_latlon=point,
            status=ConfidenceStatus.CONFIRMED,
            run_id=run_id,
            source_folder="catalog.xlsx",
        )
    )
