"""The conflict-resolution workbook: Excel as the merge tool.

When a semantic merge leaves genuine disagreements, neither copy is touched.
Instead a separate workbook is generated where a person resolves the conflicts
in the tool they already use — no JSON, no ``git mergetool``.

Three sheets:

``Info``
    Provenance: which artifact, which run, which machines, which hashes. An old
    merge workbook found months later still explains itself.

``Merge``
    The reconstructed content with every automatic merge already applied. Only
    the cells that genuinely conflict are shaded, each carrying a note with the
    base, local and Drive values. Shading one cell rather than the whole row
    keeps the eye on the actual decision.

``Conflicts``
    One row per conflicting field, with a validated ``Resolution Choice``
    dropdown — ``LOCAL``, ``DRIVE``, ``BASE`` or ``CUSTOM``.

Nothing is inferred from the file having been opened or saved: every conflict
needs an explicit choice before the merge can be applied.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from photoarchive.merge.threeway import ConflictKind, FieldConflict, MergeResult

LOG = logging.getLogger(__name__)

INFO_SHEET = "Info"
MERGE_SHEET = "Merge"
CONFLICTS_SHEET = "Conflicts"

CONFLICT_DIRECTORY = "_conflicts"

#: Allowed values of ``Resolution Choice``.
RESOLUTION_LOCAL = "LOCAL"
RESOLUTION_DRIVE = "DRIVE"
RESOLUTION_BASE = "BASE"
RESOLUTION_CUSTOM = "CUSTOM"
RESOLUTIONS: tuple[str, ...] = (
    RESOLUTION_LOCAL,
    RESOLUTION_DRIVE,
    RESOLUTION_BASE,
    RESOLUTION_CUSTOM,
)
#: A first-sync conflict has no BASE to offer — there is no common ancestor,
#: so choosing it would resurrect a value nobody agreed ever existed.
RESOLUTIONS_FIRST_SYNC: tuple[str, ...] = (
    RESOLUTION_LOCAL,
    RESOLUTION_DRIVE,
    RESOLUTION_CUSTOM,
)

#: Text shown in the ``Base`` column and in a conflict's comment when there is
#: no common baseline, so it reads as "nothing to compare against" rather than
#: as a genuinely blank base value.
NO_COMMON_BASELINE = "— no common baseline —"

CONFLICT_COLUMNS: tuple[str, ...] = (
    "Artifact",
    "Sheet",
    "Record ID",
    "Reference",
    "Field",
    "Base",
    "This computer",
    "Google Drive",
    "Resolution Choice",
    "Custom Value",
    "Resolved",
)

_HEADER_FILL = PatternFill("solid", fgColor="DDE5F0")
#: Soft red: strong enough to find, readable enough to work in.
_CONFLICT_FILL = PatternFill("solid", fgColor="FFC7CE")
_INFO_FILL = PatternFill("solid", fgColor="F2F2F2")


@dataclass(frozen=True, slots=True)
class ConflictProvenance:
    """Everything needed to understand a merge workbook later."""

    artifact_path: str
    artifact_kind: str
    run_id: str = ""
    created_at: str = ""
    machine_label: str = ""
    app_commit: str = ""
    last_sync_at: str = ""
    last_sync_machine: str = ""
    last_sync_run: str = ""
    last_sync_commit: str = ""
    base_hash: str = ""
    local_hash: str = ""
    remote_hash: str = ""
    drive_file_id: str = ""
    #: No common baseline exists yet for this artifact (see
    #: :func:`~photoarchive.merge.threeway.merge_first_sync`). Changes the
    #: Info banner and the resolution choices offered on the Conflicts sheet.
    first_sync: bool = False

    def rows(self) -> list[tuple[str, str]]:
        return [
            ("artifact path", self.artifact_path),
            ("artifact kind", self.artifact_kind),
            ("conflict run id", self.run_id),
            ("created at (UTC)", self.created_at),
            ("machine", self.machine_label),
            ("app commit", self.app_commit),
            ("", ""),
            (
                "last common sync at",
                "never — no common baseline" if self.first_sync else self.last_sync_at,
            ),
            ("last common sync machine", self.last_sync_machine),
            ("last common sync run", self.last_sync_run),
            ("last common sync commit", self.last_sync_commit),
            ("", ""),
            ("base hash", "n/a — no common baseline" if self.first_sync else self.base_hash),
            ("local hash", self.local_hash),
            ("Google Drive hash", self.remote_hash),
            ("Google Drive file id", self.drive_file_id),
        ]


def conflict_workbook_path(
    output_dir: Path | str, run_id: str, label: str, artifact: str
) -> Path:
    """Where a merge workbook goes: clearly apart from the canonical files."""
    safe = "".join("_" if character in '/\\:*?"<>|' else character for character in label)
    return (
        Path(output_dir) / CONFLICT_DIRECTORY / run_id / f"{safe}__{artifact}.merge.xlsx"
    )


def write_conflict_workbook(
    path: Path | str,
    result: MergeResult,
    provenance: ConflictProvenance,
) -> Path:
    """Generate the merge workbook for an unresolved semantic merge."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_info(workbook, provenance, result)
    _write_merge(workbook, result)
    _write_conflicts(workbook, result, provenance)

    workbook.save(path)
    workbook.close()
    LOG.info("Wrote conflict workbook %s (%s conflicts)", path, len(result.conflicts))
    return path


def _write_info(workbook: Workbook, provenance: ConflictProvenance, result: MergeResult) -> None:
    sheet = workbook.create_sheet(INFO_SHEET)
    if provenance.first_sync:
        sheet["A1"] = "FIRST SYNC — NO COMMON BASELINE"
        sheet["A1"].font = Font(bold=True, size=13, color="9C0006")
        sheet["A2"] = (
            "This computer and Google Drive have never synchronised this file before, "
            "so there is no shared ancestor to compare against — only LOCAL, DRIVE or "
            "CUSTOM can be chosen below; BASE is not offered. "
            "Neither copy has been changed. Choose a resolution for every conflict, "
            "save, then run the resolve command."
        )
    else:
        sheet["A1"] = "Merge workbook — resolve the Conflicts sheet, then apply it"
        sheet["A1"].font = Font(bold=True, size=13)
        sheet["A2"] = (
            "Neither your copy nor the Google Drive copy has been changed. "
            "Choose a resolution for every conflict, save, then run the resolve command."
        )
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    row = 4
    for label, value in provenance.rows():
        sheet.cell(row=row, column=1, value=label).font = Font(bold=bool(label))
        sheet.cell(row=row, column=1).fill = _INFO_FILL if label else PatternFill()
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    for label, value in (
        ("conflicts", len(result.conflicts)),
        ("records merged automatically", len(result.auto_merged)),
        ("records added locally", len(result.added_local)),
        ("records added on Drive", len(result.added_remote)),
    ):
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 64
    sheet.sheet_state = "visible"


def _write_merge(workbook: Workbook, result: MergeResult) -> None:
    """The reconstructed content, with only conflicting cells highlighted."""
    sheet = workbook.create_sheet(MERGE_SHEET)
    fields = _merge_fields(result)
    headers = ["Record ID", "Sheet", "Reference", *fields]

    for index, name in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL

    for offset, record_id in enumerate(result.order, start=2):
        record = result.records[record_id]
        sheet.cell(row=offset, column=1, value=record_id)
        sheet.cell(row=offset, column=2, value=record.sheet)
        sheet.cell(row=offset, column=3, value=record.label)

        for index, name in enumerate(fields, start=4):
            cell = sheet.cell(row=offset, column=index, value=record.value(name))
            conflict = result.conflict_for(record_id, name)
            if conflict is None:
                continue
            # Only this cell — colouring the row would hide which value is at stake.
            cell.fill = _CONFLICT_FILL
            cell.comment = Comment(_conflict_note(conflict), "family-photo-archive")

    sheet.freeze_panes = "D2"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 26
    for index in range(4, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 26


def _conflict_note(conflict: FieldConflict) -> str:
    if conflict.kind is ConflictKind.FIRST_SYNC:
        return (
            "FIRST SYNC — NO COMMON BASELINE\n\n"
            f"Field: {conflict.field_name}\n\n"
            f"Base:\n  {NO_COMMON_BASELINE}\n\n"
            f"This computer:\n{conflict.local or '(empty)'}\n\n"
            f"Google Drive:\n{conflict.remote or '(empty)'}\n\n"
            "Choose LOCAL, DRIVE or CUSTOM on the Conflicts sheet — BASE is not available."
        )
    return (
        "CONFLICT\n\n"
        f"Field: {conflict.field_name}\n\n"
        f"Base:\n{conflict.base or '(empty)'}\n\n"
        f"This computer:\n{conflict.local or '(empty)'}\n\n"
        f"Google Drive:\n{conflict.remote or '(empty)'}\n\n"
        "Resolve it on the Conflicts sheet."
    )


def _write_conflicts(
    workbook: Workbook, result: MergeResult, provenance: ConflictProvenance
) -> None:
    sheet = workbook.create_sheet(CONFLICTS_SHEET)
    for index, name in enumerate(CONFLICT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL

    column = get_column_letter(CONFLICT_COLUMNS.index("Resolution Choice") + 1)
    first_sync_rows: list[int] = []
    ordinary_rows: list[int] = []

    for offset, conflict in enumerate(result.conflicts, start=2):
        is_first_sync = conflict.kind is ConflictKind.FIRST_SYNC
        (first_sync_rows if is_first_sync else ordinary_rows).append(offset)
        values = {
            "Artifact": provenance.artifact_kind,
            "Sheet": conflict.sheet,
            "Record ID": conflict.record_id,
            "Reference": conflict.label,
            "Field": conflict.field_name,
            "Base": NO_COMMON_BASELINE if is_first_sync else conflict.base,
            "This computer": conflict.local,
            "Google Drive": conflict.remote,
            "Resolution Choice": "",
            "Custom Value": "",
            "Resolved": "",
        }
        for index, name in enumerate(CONFLICT_COLUMNS, start=1):
            cell = sheet.cell(row=offset, column=index, value=values[name])
            cell.alignment = Alignment(vertical="top", wrap_text=name in {
                "Base", "This computer", "Google Drive", "Custom Value"
            })
            if name == "Resolution Choice":
                cell.fill = _CONFLICT_FILL

    # Two dropdowns, not one: a first-sync row must never be offered BASE,
    # since choosing it would resurrect a value nobody agreed ever existed.
    if ordinary_rows:
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(RESOLUTIONS)}"',
            allow_blank=True,
            showDropDown=False,
        )
        validation.error = "Choose LOCAL, DRIVE, BASE or CUSTOM."
        validation.errorTitle = "Unrecognised resolution"
        sheet.add_data_validation(validation)
        for row in ordinary_rows:
            validation.add(f"{column}{row}")

    if first_sync_rows:
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(RESOLUTIONS_FIRST_SYNC)}"',
            allow_blank=True,
            showDropDown=False,
        )
        validation.error = "No common baseline: choose LOCAL, DRIVE or CUSTOM."
        validation.errorTitle = "Unrecognised resolution"
        sheet.add_data_validation(validation)
        for row in first_sync_rows:
            validation.add(f"{column}{row}")

    sheet.freeze_panes = "A2"
    widths = {
        "Artifact": 12, "Sheet": 12, "Record ID": 24, "Reference": 26, "Field": 18,
        "Base": 30, "This computer": 30, "Google Drive": 30,
        "Resolution Choice": 20, "Custom Value": 30, "Resolved": 12,
    }
    for index, name in enumerate(CONFLICT_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths[name]


def _merge_fields(result: MergeResult) -> list[str]:
    names: list[str] = []
    for record_id in result.order:
        for name in result.records[record_id].fields:
            if name not in names:
                names.append(name)
    return names


# -- Reading a resolved workbook back --------------------------------------


@dataclass(frozen=True, slots=True)
class Resolution:
    """One conflict as a person settled it."""

    record_id: str
    field_name: str
    choice: str
    custom_value: str = ""

    def value(self, conflict: FieldConflict) -> str:
        if self.choice == RESOLUTION_LOCAL:
            return conflict.local
        if self.choice == RESOLUTION_DRIVE:
            return conflict.remote
        if self.choice == RESOLUTION_BASE:
            return conflict.base
        return self.custom_value


@dataclass(slots=True)
class ResolutionSheet:
    """Everything read back from a merge workbook."""

    conflicts: list[FieldConflict]
    resolutions: dict[str, Resolution]
    provenance: dict[str, str]

    def missing(self) -> list[FieldConflict]:
        """Conflicts with no valid explicit resolution.

        Opening or saving the workbook resolves nothing: a person must choose,
        even when the choice is "keep what I already had". BASE is rejected
        for a first-sync conflict even if somebody typed it in by hand — the
        dropdown does not offer it, and there is no base value to apply.
        """
        unresolved: list[FieldConflict] = []
        for conflict in self.conflicts:
            resolution = self.resolutions.get(conflict.key)
            if resolution is None or resolution.choice not in RESOLUTIONS:
                unresolved.append(conflict)
            elif resolution.choice == RESOLUTION_CUSTOM and not resolution.custom_value:
                unresolved.append(conflict)
            elif (
                resolution.choice == RESOLUTION_BASE
                and conflict.kind is ConflictKind.FIRST_SYNC
            ):
                unresolved.append(conflict)
        return unresolved


def read_conflict_workbook(path: Path | str) -> ResolutionSheet:
    """Read the conflicts and whatever resolutions a person recorded."""
    workbook = load_workbook(Path(path))
    try:
        sheet = workbook[CONFLICTS_SHEET]
        rows = list(sheet.iter_rows(values_only=True))
        headers = {str(name): index for index, name in enumerate(rows[0]) if name}

        conflicts: list[FieldConflict] = []
        resolutions: dict[str, Resolution] = {}

        for values in rows[1:]:
            record_id = _cell(values, headers, "Record ID")
            field_name = _cell(values, headers, "Field")
            if not record_id or not field_name:
                continue

            raw_base = _cell(values, headers, "Base")
            is_first_sync = raw_base == NO_COMMON_BASELINE
            conflict = FieldConflict(
                record_id=record_id,
                label=_cell(values, headers, "Reference"),
                sheet=_cell(values, headers, "Sheet"),
                field_name=field_name,
                base="" if is_first_sync else raw_base,
                local=_cell(values, headers, "This computer"),
                remote=_cell(values, headers, "Google Drive"),
                kind=ConflictKind.FIRST_SYNC if is_first_sync else ConflictKind.FIELD,
            )
            conflicts.append(conflict)

            choice = _cell(values, headers, "Resolution Choice").upper()
            if choice:
                resolutions[conflict.key] = Resolution(
                    record_id=record_id,
                    field_name=field_name,
                    choice=choice,
                    custom_value=_cell(values, headers, "Custom Value"),
                )

        provenance = _read_info(workbook)
    finally:
        workbook.close()

    return ResolutionSheet(
        conflicts=conflicts, resolutions=resolutions, provenance=provenance
    )


def _read_info(workbook) -> dict[str, str]:
    if INFO_SHEET not in workbook.sheetnames:
        return {}
    info: dict[str, str] = {}
    for row in workbook[INFO_SHEET].iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        info[str(row[0])] = "" if len(row) < 2 or row[1] is None else str(row[1])
    return info


def _cell(values: tuple[Any, ...], headers: dict[str, int], name: str) -> str:
    index = headers.get(name)
    if index is None or index >= len(values) or values[index] is None:
        return ""
    return str(values[index]).strip()
