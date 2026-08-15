"""Rejecting a proposal: saying "this is wrong, stop offering it".

Promotion already worked; refusal did not. The rules under test:

* a rejection is a **positive statement** in its own column — an emptied cell
  means nothing, because `catalog.xlsx` is three-way merged and a missing value
  may simply be a copy that never saw it;
* a rejection is durable: no later ``learn`` revives it;
* it is reversible, because the column an alias sits in is the decision;
* an alias listed in two columns is a slip, and rejection wins.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from photoarchive.catalog.importer import import_catalog
from photoarchive.catalog.matching import find_matches
from photoarchive.catalog.models import ConfidenceStatus, EntityType
from photoarchive.catalog.service import CatalogService
from photoarchive.catalog.store import DictionaryStore

CATALOG = "catalog.xlsx"


def _store(tmp_path: Path) -> DictionaryStore:
    store = DictionaryStore(tmp_path / "dictionary.sqlite")
    store.initialize()
    return store


def _with_candidate(tmp_path: Path) -> tuple[DictionaryStore, str]:
    """A person whose dictionary holds one candidate spelling: ``мама``."""
    store = _store(tmp_path)
    person = store.add_person("Антонина Мамаева")
    store.add_alias(EntityType.PERSON, person, "мама", ConfidenceStatus.CANDIDATE)
    return store, person


def _people_row(path: Path) -> dict[str, str]:
    workbook = load_workbook(path)
    try:
        sheet = workbook["People"]
        headers = [cell.value for cell in sheet[1]]
        values = [cell.value for cell in sheet[2]]
        return {
            str(name): "" if value is None else str(value)
            for name, value in zip(headers, values)
        }
    finally:
        workbook.close()


def _set_cell(path: Path, column: str, value: str) -> None:
    """Edit one People cell, the way a person would in Excel."""
    workbook = load_workbook(path)
    try:
        sheet = workbook["People"]
        headers = [cell.value for cell in sheet[1]]
        sheet.cell(row=2, column=headers.index(column) + 1, value=value)
        workbook.save(path)
    finally:
        workbook.close()


def _export(store: DictionaryStore, tmp_path: Path) -> Path:
    path, _ = CatalogService().export(store, tmp_path / "out")
    return path


# -- The workbook surface -------------------------------------------------


def test_export_shows_rejected_aliases_in_their_own_column(tmp_path: Path) -> None:
    store, person = _with_candidate(tmp_path)
    store.set_alias_status(
        EntityType.PERSON, person, "мама", ConfidenceStatus.REJECTED
    )

    row = _people_row(_export(store, tmp_path))

    assert row["rejected_aliases"] == "мама"
    # It left the candidate column rather than appearing in both.
    assert row["candidate_aliases"] == ""


def test_rejected_alias_stays_visible_so_it_can_be_undone(tmp_path: Path) -> None:
    store, person = _with_candidate(tmp_path)
    store.set_alias_status(
        EntityType.PERSON, person, "мама", ConfidenceStatus.REJECTED
    )

    path = _export(store, tmp_path)

    assert "мама" in _people_row(path)["rejected_aliases"]


# -- Importing a decision -------------------------------------------------


def test_moving_an_alias_into_the_rejected_column_rejects_it(tmp_path: Path) -> None:
    store, _ = _with_candidate(tmp_path)
    path = _export(store, tmp_path)
    _set_cell(path, "candidate_aliases", "")
    _set_cell(path, "rejected_aliases", "мама")

    outcome = import_catalog(store, path, run_id="test")

    person = store.load().people[0]
    assert person.rejected_aliases == ("мама",)
    assert person.candidate_aliases == ()
    assert outcome.aliases_rejected == ["мама -> Антонина Мамаева"]


def test_clearing_a_candidate_cell_is_not_a_rejection(tmp_path: Path) -> None:
    store, _ = _with_candidate(tmp_path)
    path = _export(store, tmp_path)
    _set_cell(path, "candidate_aliases", "")

    import_catalog(store, path, run_id="test")

    # Absence is not a decision: another machine's copy may never have seen it.
    person = store.load().people[0]
    assert person.candidate_aliases == ("мама",)
    assert person.rejected_aliases == ()


def test_a_rejection_is_reversible(tmp_path: Path) -> None:
    store, person = _with_candidate(tmp_path)
    store.set_alias_status(
        EntityType.PERSON, person, "мама", ConfidenceStatus.REJECTED
    )
    path = _export(store, tmp_path)
    _set_cell(path, "rejected_aliases", "")
    _set_cell(path, "confirmed_aliases", "мама")

    import_catalog(store, path, run_id="test")

    loaded = store.load().people[0]
    assert loaded.confirmed_aliases == ("мама",)
    assert loaded.rejected_aliases == ()


def test_one_alias_in_two_columns_keeps_the_rejection_and_reports_it(
    tmp_path: Path,
) -> None:
    store, _ = _with_candidate(tmp_path)
    path = _export(store, tmp_path)
    _set_cell(path, "confirmed_aliases", "мама")
    _set_cell(path, "rejected_aliases", "мама")

    outcome = import_catalog(store, path, run_id="test")

    person = store.load().people[0]
    assert person.rejected_aliases == ("мама",)
    assert person.confirmed_aliases == ()
    assert outcome.collisions == ["мама -> Антонина Мамаева"]


# -- Durability -----------------------------------------------------------


def test_matching_skips_a_rejected_alias(tmp_path: Path) -> None:
    store, person = _with_candidate(tmp_path)
    store.set_alias_status(
        EntityType.PERSON, person, "мама", ConfidenceStatus.REJECTED
    )

    matches = find_matches("мама у дома", store.load(), EntityType.PERSON)

    assert matches == []


def test_a_later_pass_never_revives_a_rejected_alias(tmp_path: Path) -> None:
    """The learn → export → import cycle must not undo the decision."""
    store, person = _with_candidate(tmp_path)
    store.set_alias_status(
        EntityType.PERSON, person, "мама", ConfidenceStatus.REJECTED
    )

    # Whatever a later machine pass proposes, the refusal stands.
    store.add_alias(EntityType.PERSON, person, "мама", ConfidenceStatus.CANDIDATE)
    store.add_alias(EntityType.PERSON, person, "мама", ConfidenceStatus.CONFIRMED)

    path = _export(store, tmp_path)
    import_catalog(store, path, run_id="test")

    loaded = store.load().people[0]
    assert loaded.rejected_aliases == ("мама",)
    assert loaded.candidate_aliases == ()
    assert loaded.confirmed_aliases == ()


# -- The Evidence sheet ---------------------------------------------------


def test_evidence_sheet_lists_the_reasons_behind_a_candidate(tmp_path: Path) -> None:
    """`evidence_count` says how much; this sheet says what."""
    store, _ = _with_candidate(tmp_path)
    path = _export(store, tmp_path)
    _set_cell(path, "rejected_aliases", "мама")
    import_catalog(store, path, run_id="run-42")

    workbook = load_workbook(_export(store, tmp_path))
    try:
        sheet = workbook["Evidence"]
        headers = [cell.value for cell in sheet[1]]
        rows = [
            dict(zip(headers, [cell.value for cell in row]))
            for row in sheet.iter_rows(min_row=2)
        ]
    finally:
        workbook.close()

    assert rows, "the rejection should have left a trace"
    record = rows[0]
    assert record["entity_value"] == "Антонина Мамаева"
    assert record["candidate_text"] == "мама"
    assert record["status"] == ConfidenceStatus.REJECTED.value
    assert record["run_id"] == "run-42"


def test_evidence_sheet_is_ignored_by_the_importer(tmp_path: Path) -> None:
    store, _ = _with_candidate(tmp_path)
    path = _export(store, tmp_path)

    before = store.load().people[0]
    outcome = import_catalog(store, path, run_id="test")
    after = store.load().people[0]

    # A generated sheet is read-only: importing changes nothing through it.
    assert after.candidate_aliases == before.candidate_aliases
    assert outcome.changed == 0
