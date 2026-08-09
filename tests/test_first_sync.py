"""First sync: what happens before any common baseline exists.

None of these may guess an ancestor that never existed:

* local only -> initial upload is safe
* remote only -> initial download/adopt is safe
* both exist and byte-identical -> adopt the first baseline, no transfer
* both exist, semantically equal but byte-different -> not a conflict, but
  ``last_common_hash`` cannot be recorded until local's canonical bytes are
  uploaded over Drive — no single hash is genuinely shared until they are
* both exist and genuinely differ -> a first-sync merge workbook, with no
  BASE to offer
"""

from __future__ import annotations

import copy
from pathlib import Path

from openpyxl import load_workbook

from photoarchive.merge.apply import resolve, resolve_from_workbook
from photoarchive.merge.baseline import ARTIFACT_REVIEW, SemanticBaseline, SemanticRecord
from photoarchive.merge.semantic import semantic_baselines_equal
from photoarchive.merge.threeway import ConflictKind, merge_first_sync
from photoarchive.merge.workbook import (
    CONFLICTS_SHEET,
    INFO_SHEET,
    NO_COMMON_BASELINE,
    RESOLUTION_BASE,
    RESOLUTION_CUSTOM,
    RESOLUTION_DRIVE,
    RESOLUTION_LOCAL,
    RESOLUTIONS_FIRST_SYNC,
    ConflictProvenance,
    Resolution,
    ResolutionSheet,
    write_conflict_workbook,
)
from photoarchive.portable.models import ArtifactSyncState
from photoarchive.portable.provenance import MachineIdentity, OperationProvenance
from photoarchive.portable.sync import SyncAction, decide, record_sync

_MACHINE = MachineIdentity(machine_id="7c26e8", label="Tonya MacBook")


def _baseline(**fields: str) -> SemanticBaseline:
    return SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={
            "020": SemanticRecord(
                record_id="020", label="020.jpg", sheet="Review", fields=dict(fields)
            )
        },
        order=["020"],
    )


# -- decide(): the four base-less cases -------------------------------------


def test_case_a_local_only_is_an_initial_upload() -> None:
    decision = decide("review.xlsx", "sha256:local", None)

    assert decision.action is SyncAction.ADOPT_LOCAL
    assert decision.is_first_sync
    assert not decision.is_conflict


def test_case_b_remote_only_is_an_initial_download() -> None:
    decision = decide("review.xlsx", None, "sha256:remote")

    assert decision.action is SyncAction.ADOPT_REMOTE
    assert decision.is_first_sync
    assert not decision.is_conflict


def test_case_c1_identical_bytes_with_no_baseline_adopts_without_transfer() -> None:
    decision = decide("review.xlsx", "sha256:same", "sha256:same")

    assert decision.action is SyncAction.ADOPT_BASELINE
    assert not decision.is_conflict
    assert not decision.changes_anything


def test_case_c1_stays_a_plain_adopt_even_with_semantic_info_supplied() -> None:
    # Byte-identical already proves agreement; semantic content need not be
    # consulted, and it must not upgrade this into an upload.
    same = _baseline(place="Михнево")

    decision = decide(
        "review.xlsx", "sha256:same", "sha256:same",
        local_baseline=same, remote_baseline=same,
    )

    assert decision.action is SyncAction.ADOPT_BASELINE
    assert not decision.changes_anything


def test_case_c2_byte_different_but_semantically_equal_is_not_a_conflict() -> None:
    # Same content, packaged into two different XLSX files (different zip
    # bytes, same editable cells) -- must not become a human conflict, but a
    # canonicalising transfer is still required before a hash can be shared.
    local = _baseline(place="Михнево", people="Тоня, Мама")
    remote = _baseline(place="Михнево", people="Мама; Тоня")
    assert semantic_baselines_equal(local, remote)

    decision = decide(
        "review.xlsx", "sha256:local-bytes", "sha256:remote-bytes",
        local_baseline=local, remote_baseline=remote,
    )

    assert decision.action is SyncAction.ADOPT_BASELINE_WITH_UPLOAD
    assert not decision.is_conflict
    assert decision.changes_anything  # local -> Drive upload is still needed
    assert decision.last_common_hash is None  # nothing is genuinely shared yet


def test_case_c2_baseline_is_not_advanced_before_the_upload_succeeds() -> None:
    local = _baseline(place="Михнево", people="Тоня, Мама")
    remote = _baseline(place="Михнево", people="Мама; Тоня")
    state = ArtifactSyncState(path="review.xlsx")

    decision = decide(
        "review.xlsx", "sha256:local-bytes", "sha256:remote-bytes", state,
        local_baseline=local, remote_baseline=remote,
    )

    assert decision.action is SyncAction.ADOPT_BASELINE_WITH_UPLOAD
    # Deciding alone never writes anything.
    assert state.last_common_hash is None
    assert state.last_sync is None


def test_case_c2_after_the_upload_succeeds_the_shared_hash_is_local_hash() -> None:
    state = ArtifactSyncState(path="review.xlsx")
    provenance = OperationProvenance.create(_MACHINE, "run-9", commit="abc1234")

    # The upload makes Drive's bytes equal to local's; the hash both sides now
    # genuinely hold is local's.
    record_sync(state, "sha256:local-bytes", provenance, drive_file_id="drive-1")

    assert state.last_common_hash == "sha256:local-bytes"
    assert state.last_sync is provenance
    # A follow-up decide() against the now-shared hash is a plain no-op.
    decision = decide("review.xlsx", "sha256:local-bytes", "sha256:local-bytes", state)
    assert decision.action is SyncAction.NOOP


def test_case_d_genuinely_different_content_is_a_first_sync_conflict() -> None:
    local = _baseline(place="Дача в Михнево")
    remote = _baseline(place="Дом в Михнево")
    assert not semantic_baselines_equal(local, remote)

    decision = decide(
        "review.xlsx", "sha256:local-bytes", "sha256:remote-bytes",
        local_baseline=local, remote_baseline=remote,
    )

    assert decision.action is SyncAction.FIRST_SYNC_CONFLICT
    assert decision.is_conflict
    assert decision.is_first_sync


def test_case_d_without_semantic_info_is_conservatively_a_first_sync_conflict() -> None:
    # No baseline objects supplied: cannot prove case C2, so it must not be
    # guessed. Differing hashes fall through to case D.
    decision = decide("review.xlsx", "sha256:l", "sha256:r")

    assert decision.action is SyncAction.FIRST_SYNC_CONFLICT


def test_no_state_is_written_by_deciding_alone() -> None:
    state = ArtifactSyncState(path="review.xlsx")

    decide("review.xlsx", "sha256:l", "sha256:r")
    merge_first_sync(_baseline(place="A"), _baseline(place="B"))

    assert state.last_common_hash is None
    assert state.last_sync is None
    assert state.semantic_baseline is None


# -- merge_first_sync(): structural handling ---------------------------------


def test_local_only_records_merge_automatically() -> None:
    local = _baseline(place="Local only")
    remote = SemanticBaseline(artifact=ARTIFACT_REVIEW)

    result = merge_first_sync(local, remote)

    assert not result.has_conflicts
    assert "020" in result.records
    assert result.added_local == ["020"]


def test_remote_only_records_merge_automatically() -> None:
    local = SemanticBaseline(artifact=ARTIFACT_REVIEW)
    remote = _baseline(place="Remote only")

    result = merge_first_sync(local, remote)

    assert not result.has_conflicts
    assert "020" in result.records
    assert result.added_remote == ["020"]


def test_non_overlapping_independent_additions_auto_merge() -> None:
    local = _baseline(place="Local row")
    remote = SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={"021": SemanticRecord("021", "021.jpg", "Review", {"place": "Remote row"})},
        order=["021"],
    )

    result = merge_first_sync(local, remote)

    assert not result.has_conflicts
    assert set(result.records) == {"020", "021"}


def test_same_stable_id_added_differently_is_a_first_sync_conflict() -> None:
    local = _baseline(place="Местное")
    remote = _baseline(place="Удалённое")

    result = merge_first_sync(local, remote)

    assert result.has_conflicts
    conflict = result.conflicts[0]
    assert conflict.kind is ConflictKind.FIRST_SYNC
    assert conflict.base == ""


def test_same_stable_id_with_semantically_equal_content_is_not_a_conflict() -> None:
    local = _baseline(people="Тоня, Мама")
    remote = _baseline(people="Мама; Тоня")

    result = merge_first_sync(local, remote)

    assert not result.has_conflicts


def test_review_row_absence_is_not_treated_as_source_deletion() -> None:
    local = _baseline(place="Present locally")
    remote = SemanticBaseline(artifact=ARTIFACT_REVIEW)  # row absent on Drive

    result = merge_first_sync(local, remote, deletions_are_meaningful=False)

    assert "020" in result.records
    assert not result.has_conflicts
    assert not result.removed


# -- The first-sync merge workbook -------------------------------------------


def _first_sync_conflicting():
    local = _baseline(place="Дача в Михнево", people="Тоня")
    remote = _baseline(place="Дом в Михнево", people="Тоня")
    return merge_first_sync(local, remote)


_FIRST_SYNC_PROVENANCE = ConflictProvenance(
    artifact_path="review-output/Архив/review.xlsx",
    artifact_kind=ARTIFACT_REVIEW,
    run_id="run-1",
    machine_label="Tonya MacBook",
    local_hash="sha256:local",
    remote_hash="sha256:remote",
    first_sync=True,
)


def test_first_sync_workbook_says_no_common_baseline(tmp_path: Path) -> None:
    path = write_conflict_workbook(
        tmp_path / "m.merge.xlsx", _first_sync_conflicting(), _FIRST_SYNC_PROVENANCE
    )

    info_text = "\n".join(
        str(value)
        for row in load_workbook(path)[INFO_SHEET].iter_rows(values_only=True)
        for value in row
        if value
    )

    assert "FIRST SYNC" in info_text and "NO COMMON BASELINE" in info_text


def test_first_sync_conflicts_sheet_shows_no_common_baseline_marker(tmp_path: Path) -> None:
    path = write_conflict_workbook(
        tmp_path / "m.merge.xlsx", _first_sync_conflicting(), _FIRST_SYNC_PROVENANCE
    )

    sheet = load_workbook(path)[CONFLICTS_SHEET]
    headers = [cell.value for cell in sheet[1]]
    base_value = sheet.cell(row=2, column=headers.index("Base") + 1).value

    assert base_value == NO_COMMON_BASELINE


def test_first_sync_conflict_note_has_no_base_and_says_first_sync(tmp_path: Path) -> None:
    from photoarchive.merge.workbook import MERGE_SHEET

    path = write_conflict_workbook(
        tmp_path / "m.merge.xlsx", _first_sync_conflicting(), _FIRST_SYNC_PROVENANCE
    )

    sheet = load_workbook(path)[MERGE_SHEET]
    headers = [cell.value for cell in sheet[1]]
    note = sheet.cell(row=2, column=headers.index("place") + 1).comment.text

    assert "FIRST SYNC" in note and "NO COMMON BASELINE" in note
    assert "Дача в Михнево" in note and "Дом в Михнево" in note


def test_base_resolution_is_rejected_for_a_first_sync_conflict() -> None:
    result = _first_sync_conflicting()
    key = result.conflicts[0].key
    sheet = ResolutionSheet(
        conflicts=result.conflicts,
        resolutions={
            key: Resolution(
                record_id=key.split("::")[0], field_name=key.split("::")[1],
                choice=RESOLUTION_BASE,
            )
        },
        provenance={},
    )

    outcome = resolve(result, sheet)

    assert not outcome.ok
    assert outcome.unresolved


def test_local_drive_and_custom_still_work_for_a_first_sync_conflict() -> None:
    result = _first_sync_conflicting()
    key = result.conflicts[0].key

    for choice, custom, expected in (
        (RESOLUTION_LOCAL, "", "Дача в Михнево"),
        (RESOLUTION_DRIVE, "", "Дом в Михнево"),
        (RESOLUTION_CUSTOM, "Михнево, дача", "Михнево, дача"),
    ):
        sheet = ResolutionSheet(
            conflicts=result.conflicts,
            resolutions={
                key: Resolution(
                    record_id=key.split("::")[0], field_name=key.split("::")[1],
                    choice=choice, custom_value=custom,
                )
            },
            provenance={},
        )
        outcome = resolve(result, sheet)

        assert outcome.ok
        assert outcome.records["020"].value("place") == expected


def test_resolution_dropdown_omits_base_for_first_sync() -> None:
    assert RESOLUTION_BASE not in RESOLUTIONS_FIRST_SYNC
    assert set(RESOLUTIONS_FIRST_SYNC) == {RESOLUTION_LOCAL, RESOLUTION_DRIVE, RESOLUTION_CUSTOM}


def test_a_first_sync_workbook_round_trips_through_excel(tmp_path: Path) -> None:
    result = _first_sync_conflicting()
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", result, _FIRST_SYNC_PROVENANCE)

    workbook = load_workbook(path)
    sheet = workbook[CONFLICTS_SHEET]
    headers = [cell.value for cell in sheet[1]]
    sheet.cell(row=2, column=headers.index("Resolution Choice") + 1).value = RESOLUTION_LOCAL
    workbook.save(path)

    outcome = resolve_from_workbook(result, path)

    assert outcome.ok
    assert outcome.records["020"].value("place") == "Дача в Михнево"
