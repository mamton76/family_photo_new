"""Dictionary matching, suggestions, learning, evidence and coordinates."""

from __future__ import annotations

from pathlib import Path

from photoarchive.catalog.learning import LearningContext, learn_from_rows
from photoarchive.catalog.matching import (
    confirmed_canonicals,
    find_matches,
)
from photoarchive.catalog.models import ConfidenceStatus, EntityType
from photoarchive.catalog.store import DictionaryStore
from photoarchive.geo import LatLon, parse_latlon
from photoarchive.models import WorkflowStatus
from photoarchive.parsing.suggestions import suggest, suggest_date
from photoarchive.review.model import ReviewRow

MOSCOW = LatLon(55.751244, 37.618423)
FAR_AWAY = LatLon(59.934280, 30.335099)


def _store(tmp_path: Path) -> DictionaryStore:
    store = DictionaryStore(tmp_path / "dictionary.sqlite")
    store.initialize()
    return store


def _seeded(tmp_path: Path) -> DictionaryStore:
    """A store with the canonical example from the specification."""
    store = _store(tmp_path)
    person = store.add_person("Антонина Мамаева")
    store.add_alias(EntityType.PERSON, person, "Тоня Мамаева", ConfidenceStatus.CONFIRMED)
    store.add_alias(EntityType.PERSON, person, "Тоня", ConfidenceStatus.CONFIRMED)
    return store


# -- Matching -------------------------------------------------------------


def test_exact_canonical_match(tmp_path: Path) -> None:
    store = _seeded(tmp_path)

    matches = find_matches("Антонина Мамаева дома", store.load(), EntityType.PERSON)

    assert confirmed_canonicals(matches) == ["Антонина Мамаева"]


def test_confirmed_alias_match(tmp_path: Path) -> None:
    store = _seeded(tmp_path)

    matches = find_matches("Тоня Мамаева (3г)", store.load(), EntityType.PERSON)

    assert confirmed_canonicals(matches) == ["Антонина Мамаева"]


def test_longer_phrase_beats_a_shorter_alias(tmp_path: Path) -> None:
    store = _seeded(tmp_path)

    matches = find_matches("Тоня Мамаева (3г)", store.load(), EntityType.PERSON)

    # "Тоня Мамаева" claims the span, so bare "Тоня" is not reported again.
    assert len(matches) == 1
    assert matches[0].matched_text == "Тоня Мамаева"


def test_matching_respects_word_boundaries(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_person("Ян")

    assert find_matches("Январь был холодный", store.load(), EntityType.PERSON) == []


def test_candidate_alias_is_not_a_confirmed_fact(tmp_path: Path) -> None:
    store = _store(tmp_path)
    person = store.add_person("Антонина Мамаева")
    store.add_alias(EntityType.PERSON, person, "Тонечка", ConfidenceStatus.CANDIDATE)

    matches = find_matches("Тонечка дома", store.load(), EntityType.PERSON)

    assert len(matches) == 1
    assert not matches[0].is_confirmed
    assert confirmed_canonicals(matches) == []


def test_candidate_alias_never_reaches_a_suggestion(tmp_path: Path) -> None:
    store = _store(tmp_path)
    person = store.add_person("Антонина Мамаева")
    store.add_alias(EntityType.PERSON, person, "Тонечка", ConfidenceStatus.CANDIDATE)

    suggestion = suggest("Тонечка дома", store.load())

    assert suggestion.people == ()
    assert len(suggestion.candidates) == 1


# -- Suggestions ----------------------------------------------------------


def test_suggested_people_use_canonical_names(tmp_path: Path) -> None:
    store = _seeded(tmp_path)

    suggestion = suggest("1979. Тоня Мамаева (3г). Дома.", store.load())

    assert suggestion.people == ("Антонина Мамаева",)
    assert suggestion.people_matched == 1


def test_suggested_date_reads_year_month_and_day() -> None:
    assert suggest_date("1979. Тоня Мамаева (3г).") == "1979"
    assert suggest_date("1980 июнь. В зоопарке.") == "1980-06"
    assert suggest_date("1979 февраль 19. Тоня Мамаева.") == "1979-02-19"


def test_no_year_means_no_date_suggestion() -> None:
    assert suggest_date("Тоня Мамаева (3г). Дома.") == ""
    assert suggest_date("") == ""


def test_ages_never_become_birth_years() -> None:
    # "(2.5г)" plus a date must not produce a derived year anywhere.
    assert suggest_date("1979 февраль 19. Тоня Мамаева (2.5г).") == "1979-02-19"


def test_empty_dictionary_yields_empty_entity_suggestions(tmp_path: Path) -> None:
    suggestion = suggest("1979. Тоня Мамаева. Дома на Днепропетровской.", _store(tmp_path).load())

    assert suggestion.people == ()
    assert suggestion.place == ""
    assert suggestion.latlon == ""
    assert suggestion.date == "1979"


def test_confirmed_place_coordinates_populate_suggested_latlon(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)

    suggestion = suggest("1979 май. На даче в Михнево.", store.load())

    assert suggestion.place == "Михнево"
    assert suggestion.latlon == MOSCOW.format()
    assert suggestion.coordinates_reused == 1


def test_place_without_confirmed_coordinates_suggests_none(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево")

    suggestion = suggest("На даче в Михнево.", store.load())

    assert suggestion.place == "Михнево"
    assert suggestion.latlon == ""


def test_section_context_contributes_people(tmp_path: Path) -> None:
    store = _seeded(tmp_path)

    suggestion = suggest("У кого-то в гостях.", store.load(), "Далее Тоня")

    assert suggestion.people == ("Антонина Мамаева",)


# -- Learning -------------------------------------------------------------


def _approved(**kwargs) -> ReviewRow:
    row = ReviewRow(reference=kwargs.pop("reference", "020"), **kwargs)
    row.status = WorkflowStatus.APPROVED
    return row


def test_human_values_teach_regardless_of_workflow_status(tmp_path: Path) -> None:
    # APPROVED is a publication state, not a precondition for learning: a name
    # typed into People on a NEW row is just as much a stated fact.
    store = _store(tmp_path)
    pending = ReviewRow(reference="020", people="Антонина Мамаева")
    pending.status = WorkflowStatus.NEW

    outcome = learn_from_rows(store, [pending])

    assert outcome.new_people == ["Антонина Мамаева"]


def test_rows_with_no_human_values_teach_nothing(tmp_path: Path) -> None:
    store = _store(tmp_path)
    empty = ReviewRow(reference="020", suggested_people="Только Предположение")
    empty.status = WorkflowStatus.NEW

    outcome = learn_from_rows(store, [empty])

    assert outcome.new_people == []
    assert store.load().people == []


def test_approved_metadata_creates_a_new_canonical_person(tmp_path: Path) -> None:
    store = _store(tmp_path)

    outcome = learn_from_rows(store, [_approved(people="Антонина Мамаева")])

    assert outcome.new_people == ["Антонина Мамаева"]
    assert [person.canonical_name for person in store.load().people] == ["Антонина Мамаева"]


def test_approved_row_creates_places_and_tags(tmp_path: Path) -> None:
    store = _store(tmp_path)

    outcome = learn_from_rows(store, [_approved(place="Михнево", tags="дача")])

    assert outcome.new_places == ["Михнево"]
    assert outcome.new_tags == ["дача"]


def test_unambiguous_correction_adds_a_confirmed_alias(tmp_path: Path) -> None:
    store = _store(tmp_path)
    row = _approved(people="Антонина Мамаева", source_description="Тоня Мамаева")

    outcome = learn_from_rows(store, [row])

    person = store.load().people[0]
    assert "Тоня Мамаева" in person.confirmed_aliases
    assert outcome.confirmed_aliases == ["Тоня Мамаева -> Антонина Мамаева"]


def test_ambiguous_mapping_creates_candidates_not_confirmed_aliases(tmp_path: Path) -> None:
    store = _store(tmp_path)
    row = _approved(
        people="Антонина Мамаева",
        source_description="Тоня Мамаева. Сережа Мамаев. Настя Платова.",
    )

    outcome = learn_from_rows(store, [row])

    person = store.load().people[0]
    assert person.confirmed_aliases == ()
    assert len(person.candidate_aliases) == 3
    assert outcome.confirmed_aliases == []
    assert len(outcome.candidate_aliases) == 3


def test_kinship_terms_do_not_become_universal_aliases(tmp_path: Path) -> None:
    store = _store(tmp_path)
    row = _approved(people="Антонина Мамаева", source_description="мама")

    learn_from_rows(store, [row])

    person = store.load().people[0]
    assert person.confirmed_aliases == ()
    assert "мама" in person.candidate_aliases


def test_evidence_is_recorded_and_kept(tmp_path: Path) -> None:
    store = _store(tmp_path)
    row = _approved(people="Антонина Мамаева", source_description="Тоня Мамаева")

    learn_from_rows(
        store, [row], LearningContext(source_root="root-1", source_folder="/", run_id="run-1")
    )

    evidence = store.evidence_for(EntityType.PERSON, "Антонина Мамаева")
    assert len(evidence) >= 1
    assert any(item.candidate_text == "Тоня Мамаева" for item in evidence)
    assert evidence[0].run_id == "run-1"
    assert evidence[0].reference == "020"


def test_evidence_survives_promotion(tmp_path: Path) -> None:
    store = _store(tmp_path)
    ambiguous = _approved(
        people="Антонина Мамаева", source_description="Тоня Мамаева. Сережа Мамаев."
    )
    learn_from_rows(store, [ambiguous])
    before = store.evidence_count(EntityType.PERSON, "Антонина Мамаева")

    person_id = store.load().people[0].person_id
    store.add_alias(
        EntityType.PERSON, person_id, "Тоня Мамаева", ConfidenceStatus.CONFIRMED
    )

    after = store.evidence_count(EntityType.PERSON, "Антонина Мамаева")
    assert after >= before
    assert "Тоня Мамаева" in store.load().people[0].confirmed_aliases


def test_confirmed_alias_is_not_downgraded_to_candidate(tmp_path: Path) -> None:
    store = _seeded(tmp_path)
    person_id = store.load().people[0].person_id

    store.add_alias(EntityType.PERSON, person_id, "Тоня", ConfidenceStatus.CANDIDATE)

    assert "Тоня" in store.load().people[0].confirmed_aliases


# -- Place coordinates ----------------------------------------------------


def test_approved_coordinates_are_stored_for_a_new_place(tmp_path: Path) -> None:
    store = _store(tmp_path)
    row = _approved(place="Михнево", latlon=MOSCOW.format())

    outcome = learn_from_rows(store, [row])

    place = store.load().places[0]
    assert place.latlon is not None
    assert place.latlon.format() == MOSCOW.format()
    assert outcome.latlon_added == ["Михнево"]


def test_conflicting_coordinates_do_not_overwrite_confirmed_ones(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)

    outcome = learn_from_rows(
        store, [_approved(place="Михнево", latlon=FAR_AWAY.format())]
    )

    place = store.load().places[0]
    assert place.latlon.format() == MOSCOW.format()
    assert [point.format() for point in place.candidate_latlon] == [FAR_AWAY.format()]
    assert outcome.latlon_conflicts == ["Михнево"]
    assert outcome.latlon_added == []


def test_matching_coordinates_are_not_a_conflict(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)
    nearly = LatLon(MOSCOW.latitude + 0.000005, MOSCOW.longitude)

    outcome = learn_from_rows(store, [_approved(place="Михнево", latlon=nearly.format())])

    assert outcome.latlon_conflicts == []
    assert store.load().places[0].candidate_latlon == []


def test_candidate_coordinates_are_never_suggested(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)
    learn_from_rows(store, [_approved(place="Михнево", latlon=FAR_AWAY.format())])

    suggestion = suggest("На даче в Михнево.", store.load())

    assert suggestion.latlon == MOSCOW.format()


def test_store_is_idempotent_for_repeated_entities(tmp_path: Path) -> None:
    store = _store(tmp_path)
    first = store.add_person("Антонина Мамаева")
    second = store.add_person("Антонина Мамаева")

    assert first == second
    assert len(store.load().people) == 1


def test_place_latlon_round_trips_through_storage(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)

    place = store.load().places[0]

    assert parse_latlon(place.latlon.format()) is not None
    assert place.map_link and "55.751244" in place.map_link


# -- catalog.xlsx export --------------------------------------------------


def test_catalog_workbook_has_the_three_sheets(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from photoarchive.catalog.service import CATALOG_SHEETS, CatalogService

    store = _seeded(tmp_path)
    path, counts = CatalogService().export(store, tmp_path / "out")

    assert path.name == "catalog.xlsx"
    workbook = load_workbook(path)
    assert tuple(workbook.sheetnames) == CATALOG_SHEETS
    assert counts.people == 1


def test_catalog_shows_confirmed_and_candidate_aliases_apart(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from photoarchive.catalog.service import CatalogService

    store = _seeded(tmp_path)
    person_id = store.load().people[0].person_id
    store.add_alias(EntityType.PERSON, person_id, "Тонечка", ConfidenceStatus.CANDIDATE)

    path, counts = CatalogService().export(store, tmp_path / "out")

    sheet = load_workbook(path)["People"]
    headers = [cell.value for cell in sheet[1]]
    confirmed = sheet.cell(row=2, column=headers.index("confirmed_aliases") + 1)
    candidate = sheet.cell(row=2, column=headers.index("candidate_aliases") + 1)

    assert "Тоня Мамаева" in confirmed.value
    assert candidate.value == "Тонечка"
    assert counts.candidate_aliases == 1
    # Candidates are shaded so they cannot be mistaken for facts.
    assert candidate.fill.fgColor.rgb.endswith("FFF2CC")


def test_catalog_separates_confirmed_and_candidate_coordinates(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from photoarchive.catalog.service import CatalogService

    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)
    learn_from_rows(store, [_approved(place="Михнево", latlon=FAR_AWAY.format())])

    path, counts = CatalogService().export(store, tmp_path / "out")

    sheet = load_workbook(path)["Places"]
    headers = [cell.value for cell in sheet[1]]
    assert sheet.cell(row=2, column=headers.index("latlon") + 1).value == MOSCOW.format()
    assert sheet.cell(row=2, column=headers.index("candidate_latlon") + 1).value == FAR_AWAY.format()
    assert counts.candidate_coordinates == 1


def test_catalog_reports_evidence_counts(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from photoarchive.catalog.service import CatalogService

    store = _store(tmp_path)
    learn_from_rows(store, [_approved(people="Антонина Мамаева", source_description="Тоня")])

    path, _ = CatalogService().export(store, tmp_path / "out")

    sheet = load_workbook(path)["People"]
    headers = [cell.value for cell in sheet[1]]
    assert sheet.cell(row=2, column=headers.index("evidence_count") + 1).value >= 1


def test_empty_catalog_still_exports(tmp_path: Path) -> None:
    from photoarchive.catalog.service import CatalogService

    path, counts = CatalogService().export(_store(tmp_path), tmp_path / "out")

    assert path.exists()
    assert (counts.people, counts.places, counts.tags) == (0, 0, 0)
