"""Semantic three-way merge and the Excel conflict-resolution workbook."""

from __future__ import annotations

import copy
from pathlib import Path

from openpyxl import load_workbook

from photoarchive.merge.apply import (
    ApplyStatus,
    archive_merge_workbook,
    resolve,
    resolve_from_workbook,
)
from photoarchive.merge.baseline import (
    ARTIFACT_CATALOG,
    ARTIFACT_REVIEW,
    REVIEW_HUMAN_FIELDS,
    REVIEW_MACHINE_FIELDS,
    SemanticBaseline,
    SemanticRecord,
    baseline_from_catalog,
    baseline_from_review_rows,
)
from photoarchive.merge.semantic import semantic_equal
from photoarchive.merge.threeway import ConflictKind, merge
from photoarchive.merge.workbook import (
    CONFLICTS_SHEET,
    INFO_SHEET,
    MERGE_SHEET,
    RESOLUTION_BASE,
    RESOLUTION_CUSTOM,
    RESOLUTION_DRIVE,
    RESOLUTION_LOCAL,
    ConflictProvenance,
    Resolution,
    ResolutionSheet,
    conflict_workbook_path,
    read_conflict_workbook,
    write_conflict_workbook,
)
from photoarchive.models import WorkflowStatus
from photoarchive.review.model import ReviewRow

PROVENANCE = ConflictProvenance(
    artifact_path="review-output/Архив/review.xlsx",
    artifact_kind=ARTIFACT_REVIEW,
    run_id="run-1",
    created_at="2026-08-09T12:00:00Z",
    machine_label="Tonya MacBook",
    app_commit="df0fb26",
    last_sync_at="2026-08-08T10:42:17Z",
    last_sync_machine="Home PC",
    base_hash="sha256:base",
    local_hash="sha256:local",
    remote_hash="sha256:remote",
    drive_file_id="drive-1",
)


def _baseline(**fields: str) -> SemanticBaseline:
    return SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={
            "020": SemanticRecord(
                record_id="020", label="020.jpg", sheet="Review", fields=dict(fields)
            )
        },
        order=["020"],
    )


def _sides(**fields: str):
    base = _baseline(**fields)
    return base, copy.deepcopy(base), copy.deepcopy(base)


# -- Field-level merge -----------------------------------------------------


def test_no_changes_merges_to_the_base() -> None:
    base, local, remote = _sides(place="Михнево", people="Тоня")

    result = merge(base, local, remote)

    assert not result.has_conflicts
    assert result.records["020"].fields == {"place": "Михнево", "people": "Тоня"}


def test_local_only_edit_wins() -> None:
    base, local, remote = _sides(place="Михнево")
    local.records["020"].fields["place"] = "Дача в Михнево"

    result = merge(base, local, remote)

    assert not result.has_conflicts
    assert result.records["020"].value("place") == "Дача в Михнево"


def test_remote_only_edit_wins() -> None:
    base, local, remote = _sides(place="Михнево")
    remote.records["020"].fields["place"] = "Дом в Михнево"

    result = merge(base, local, remote)

    assert not result.has_conflicts
    assert result.records["020"].value("place") == "Дом в Михнево"


def test_independent_fields_merge_without_asking() -> None:
    # The specification's example, verbatim.
    base, local, remote = _sides(place="Михнево", people="Тоня")
    local.records["020"].fields["place"] = "Дача в Михнево"
    remote.records["020"].fields["people"] = "Тоня, Мама"

    result = merge(base, local, remote)

    assert not result.has_conflicts
    assert result.records["020"].fields == {
        "place": "Дача в Михнево",
        "people": "Тоня, Мама",
    }


def test_the_same_new_value_on_both_sides_is_not_a_conflict() -> None:
    base, local, remote = _sides(place="Михнево")
    local.records["020"].fields["place"] = "Дача в Михнево"
    remote.records["020"].fields["place"] = "Дача в Михнево"

    result = merge(base, local, remote)

    assert not result.has_conflicts
    assert result.records["020"].value("place") == "Дача в Михнево"


def test_the_same_field_changed_differently_is_a_conflict() -> None:
    base, local, remote = _sides(place="Михнево")
    local.records["020"].fields["place"] = "Дача в Михнево"
    remote.records["020"].fields["place"] = "Дом в Михнево"

    result = merge(base, local, remote)

    assert len(result.conflicts) == 1
    conflict = result.conflicts[0]
    assert (conflict.base, conflict.local, conflict.remote) == (
        "Михнево", "Дача в Михнево", "Дом в Михнево"
    )
    # Nothing is guessed: the value waits at base until a person decides.
    assert result.records["020"].value("place") == "Михнево"


def test_several_conflicts_in_one_row() -> None:
    base, local, remote = _sides(place="Михнево", people="Тоня")
    local.records["020"].fields.update(place="Дача", people="Аня")
    remote.records["020"].fields.update(place="Дом", people="Мама")

    result = merge(base, local, remote)

    assert {conflict.field_name for conflict in result.conflicts} == {"place", "people"}


def test_conflicts_in_different_rows() -> None:
    base = SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={
            "020": SemanticRecord("020", "020.jpg", "Review", {"place": "A"}),
            "021": SemanticRecord("021", "021.jpg", "Review", {"place": "B"}),
        },
        order=["020", "021"],
    )
    local, remote = copy.deepcopy(base), copy.deepcopy(base)
    local.records["020"].fields["place"] = "A-local"
    remote.records["020"].fields["place"] = "A-remote"
    local.records["021"].fields["place"] = "B-local"
    remote.records["021"].fields["place"] = "B-remote"

    result = merge(base, local, remote)

    assert {conflict.record_id for conflict in result.conflicts} == {"020", "021"}


def test_machine_owned_fields_are_absent_from_the_baseline() -> None:
    row = ReviewRow(
        reference="020", filename="020.jpg", place="Михнево",
        suggested_place="Канонично", source_description="исходный текст",
        review_reason="Description changed",
    )

    baseline = baseline_from_review_rows([row])

    fields = baseline.records["020"].fields
    assert set(fields) == set(REVIEW_HUMAN_FIELDS)
    for name in REVIEW_MACHINE_FIELDS:
        assert name not in fields


def test_machine_owned_differences_cannot_conflict() -> None:
    base = baseline_from_review_rows([ReviewRow(reference="020", place="Михнево")])
    local = baseline_from_review_rows(
        [ReviewRow(reference="020", place="Михнево", suggested_place="X")]
    )
    remote = baseline_from_review_rows(
        [ReviewRow(reference="020", place="Михнево", suggested_place="Y")]
    )

    assert not merge(base, local, remote).has_conflicts


def test_status_is_human_owned_and_survives() -> None:
    base, local, remote = _sides(status="NEW")
    local.records["020"].fields["status"] = WorkflowStatus.APPROVED.value

    result = merge(base, local, remote)

    assert result.records["020"].value("status") == "APPROVED"


# -- Structural changes ----------------------------------------------------


def test_independent_additions_merge() -> None:
    base = SemanticBaseline(artifact=ARTIFACT_REVIEW)
    local = _baseline(place="Local only")
    remote = SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={"021": SemanticRecord("021", "021.jpg", "Review", {"place": "Remote"})},
        order=["021"],
    )

    result = merge(base, local, remote)

    assert set(result.records) == {"020", "021"}
    assert not result.has_conflicts


def test_the_same_id_added_differently_conflicts() -> None:
    base = SemanticBaseline(artifact=ARTIFACT_REVIEW)
    local = _baseline(place="Местное")
    remote = _baseline(place="Удалённое")

    result = merge(base, local, remote)

    assert result.has_conflicts
    assert result.conflicts[0].kind is ConflictKind.ADDED_BOTH


def test_a_deleted_review_row_is_not_a_source_deletion() -> None:
    # Source scanning owns which rows exist; a row missing from somebody's copy
    # must never delete an archive item.
    base, local, remote = _sides(place="Михнево")
    del local.records["020"]
    local.order.remove("020")

    result = merge(base, local, remote, deletions_are_meaningful=False)

    assert "020" in result.records
    assert not result.has_conflicts


def test_a_deleted_catalog_entity_is_honoured() -> None:
    base = SemanticBaseline(
        artifact=ARTIFACT_CATALOG,
        records={"People:p1": SemanticRecord("People:p1", "Тоня", "People", {"notes": ""})},
        order=["People:p1"],
    )
    local, remote = copy.deepcopy(base), copy.deepcopy(base)
    del local.records["People:p1"]
    local.order.clear()

    result = merge(base, local, remote, deletions_are_meaningful=True)

    assert "People:p1" not in result.records
    assert not result.has_conflicts


def test_deleting_what_the_other_side_edited_is_a_conflict() -> None:
    base = SemanticBaseline(
        artifact=ARTIFACT_CATALOG,
        records={"People:p1": SemanticRecord("People:p1", "Тоня", "People", {"notes": "a"})},
        order=["People:p1"],
    )
    local, remote = copy.deepcopy(base), copy.deepcopy(base)
    del local.records["People:p1"]
    local.order.clear()
    remote.records["People:p1"].fields["notes"] = "b"

    result = merge(base, local, remote, deletions_are_meaningful=True)

    assert result.has_conflicts
    assert result.conflicts[0].kind is ConflictKind.DELETE_VS_EDIT


def test_catalog_baseline_keys_on_stable_ids(tmp_path: Path) -> None:
    from photoarchive.catalog.store import DictionaryStore
    from photoarchive.geo import LatLon

    store = DictionaryStore(tmp_path / "dict.sqlite")
    store.initialize()
    place_id = store.add_place("Михнево", LatLon(55.751244, 37.618423))

    baseline = baseline_from_catalog(store.load())

    record = baseline.records[f"Places:{place_id}"]
    assert record.value("canonical_place") == "Михнево"
    assert record.value("latlon") == "55.751244, 37.618423"


# -- Semantic normalization -------------------------------------------------


def test_people_same_members_different_separator_or_order_is_not_a_conflict() -> None:
    base, local, remote = _sides(people="Тоня")
    local.records["020"].fields["people"] = "Тоня, Мама"
    remote.records["020"].fields["people"] = "Мама; Тоня"

    result = merge(base, local, remote)

    assert not result.has_conflicts
    # Deterministic, and not a rewrite: one side's actual wording is kept.
    assert result.records["020"].value("people") in {"Тоня, Мама", "Мама; Тоня"}


def test_tags_same_members_different_separator_or_order_is_not_a_conflict() -> None:
    base, local, remote = _sides(tags="")
    local.records["020"].fields["tags"] = "школа, праздник"
    remote.records["020"].fields["tags"] = "праздник; школа"

    result = merge(base, local, remote)

    assert not result.has_conflicts


def test_albums_equivalent_formatting_is_not_a_conflict() -> None:
    base, local, remote = _sides(albums="")
    local.records["020"].fields["albums"] = "Семья; Школа"
    remote.records["020"].fields["albums"] = "Школа, Семья"

    result = merge(base, local, remote)

    assert not result.has_conflicts


def test_latlon_spacing_and_formatting_is_not_a_conflict() -> None:
    base, local, remote = _sides(latlon="")
    local.records["020"].fields["latlon"] = "55.619898,37.598040"
    remote.records["020"].fields["latlon"] = "55.619898, 37.598040"

    result = merge(base, local, remote)

    assert not result.has_conflicts


def test_two_genuinely_different_latlon_values_conflict_when_both_changed() -> None:
    base, local, remote = _sides(latlon="0,0")
    local.records["020"].fields["latlon"] = "55.619898, 37.598040"
    remote.records["020"].fields["latlon"] = "59.933333, 30.316667"

    result = merge(base, local, remote)

    assert result.has_conflicts
    assert result.conflicts[0].field_name == "latlon"


def test_blank_none_and_empty_string_are_not_a_conflict() -> None:
    base, local, remote = _sides(notes="")
    local.records["020"].fields["notes"] = None  # type: ignore[assignment]
    remote.records["020"].fields["notes"] = "   "

    result = merge(base, local, remote)

    assert not result.has_conflicts
    assert result.records["020"].value("notes") == ""


def test_semantic_equal_treats_none_empty_and_whitespace_as_blank() -> None:
    assert semantic_equal(ARTIFACT_REVIEW, "notes", None, "")
    assert semantic_equal(ARTIFACT_REVIEW, "place", "   ", "")


def test_date_whitespace_trivia_is_not_a_conflict() -> None:
    base, local, remote = _sides(date="1979-05-17")
    local.records["020"].fields["date"] = "1979-05-17"
    remote.records["020"].fields["date"] = " 1979-05-17 "

    result = merge(base, local, remote)

    assert not result.has_conflicts


def test_date_year_and_month_precision_are_different_values() -> None:
    # A year-only value is never normalised into a fake full date.
    assert not semantic_equal(ARTIFACT_REVIEW, "date", "1979", "1979-05")


def test_date_precision_difference_is_a_genuine_conflict_when_both_change() -> None:
    base, local, remote = _sides(date="1978")
    local.records["020"].fields["date"] = "1979"
    remote.records["020"].fields["date"] = "1979-05"

    result = merge(base, local, remote)

    assert result.has_conflicts
    assert result.conflicts[0].field_name == "date"


def test_place_alias_or_canonical_spelling_is_not_silently_collapsed() -> None:
    # Place is human-owned final text; a dictionary could plausibly resolve
    # both spellings to one canonical place, but that is not this field's
    # contract (that is what Suggested Place is for).
    assert not semantic_equal(ARTIFACT_REVIEW, "place", "Дача в Михнево", "Дом в Михнево")
    # Basic whitespace normalisation is still fine.
    assert semantic_equal(ARTIFACT_REVIEW, "place", "  Михнево ", "Михнево")


def test_prose_differences_remain_differences() -> None:
    base, local, remote = _sides(notes="")
    local.records["020"].fields["notes"] = "подписано на обороте"
    remote.records["020"].fields["notes"] = "возможно 1980"

    result = merge(base, local, remote)

    assert result.has_conflicts
    assert result.conflicts[0].field_name == "notes"


def test_prose_line_ending_trivia_is_not_a_conflict() -> None:
    base, local, remote = _sides(notes="line one\nline two")
    local.records["020"].fields["notes"] = "line one\nline two"
    remote.records["020"].fields["notes"] = "line one\r\nline two"

    result = merge(base, local, remote)

    assert not result.has_conflicts


def test_prose_is_not_reordered_or_whitespace_collapsed() -> None:
    assert not semantic_equal(
        ARTIFACT_REVIEW, "notes", "line one\nline two", "line two\nline one"
    )
    assert not semantic_equal(ARTIFACT_REVIEW, "notes", "a  b", "a b")


def test_status_differences_remain_differences() -> None:
    base, local, remote = _sides(status="NEW")
    local.records["020"].fields["status"] = WorkflowStatus.APPROVED.value
    remote.records["020"].fields["status"] = "REVIEW"

    result = merge(base, local, remote)

    assert result.has_conflicts
    assert result.conflicts[0].field_name == "status"


def test_catalog_confirmed_aliases_set_semantics_no_conflict() -> None:
    base = SemanticBaseline(
        artifact=ARTIFACT_CATALOG,
        records={
            "People:p1": SemanticRecord(
                "People:p1", "Тоня", "People", {"confirmed_aliases": "Antonina"}
            )
        },
        order=["People:p1"],
    )
    local, remote = copy.deepcopy(base), copy.deepcopy(base)
    local.records["People:p1"].fields["confirmed_aliases"] = "Antonina; Тонечка"
    remote.records["People:p1"].fields["confirmed_aliases"] = "Тонечка;Antonina"

    result = merge(base, local, remote, deletions_are_meaningful=True)

    assert not result.has_conflicts


def _candidate_latlon_baseline(value: str) -> SemanticBaseline:
    return SemanticBaseline(
        artifact=ARTIFACT_CATALOG,
        records={
            "Places:pl1": SemanticRecord(
                "Places:pl1", "Михнево", "Places", {"candidate_latlon": value}
            )
        },
        order=["Places:pl1"],
    )


def test_candidate_latlon_set_different_order_is_not_a_conflict() -> None:
    base = _candidate_latlon_baseline("")
    local = _candidate_latlon_baseline(
        "55.619898,37.598040; 55.620000,37.599000"
    )
    remote = _candidate_latlon_baseline(
        "55.620000,37.599000; 55.619898,37.598040"
    )

    result = merge(base, local, remote, deletions_are_meaningful=True)

    assert not result.has_conflicts


def test_candidate_latlon_set_spacing_differences_are_not_a_conflict() -> None:
    base = _candidate_latlon_baseline("")
    local = _candidate_latlon_baseline("55.619898,37.598040; 55.620000,37.599000")
    remote = _candidate_latlon_baseline("55.620000, 37.599000; 55.619898, 37.598040")

    result = merge(base, local, remote, deletions_are_meaningful=True)

    assert not result.has_conflicts


def test_candidate_latlon_set_different_coordinates_conflict() -> None:
    base = _candidate_latlon_baseline("55.619898,37.598040")
    local = _candidate_latlon_baseline("55.619898,37.598040; 55.620000,37.599000")
    remote = _candidate_latlon_baseline("55.619898,37.598040; 59.933333,30.316667")

    result = merge(base, local, remote, deletions_are_meaningful=True)

    assert result.has_conflicts
    assert result.conflicts[0].field_name == "candidate_latlon"


def test_candidate_latlon_set_invalid_values_are_not_silently_dropped() -> None:
    # An unparsable candidate is kept as its own token, not discarded and not
    # equated with a different unparsable token.
    assert not semantic_equal(
        ARTIFACT_CATALOG, "candidate_latlon",
        "not-a-coordinate; 55.619898,37.598040",
        "55.619898,37.598040",
    )
    assert not semantic_equal(
        ARTIFACT_CATALOG, "candidate_latlon", "not-a-coordinate", "also-not-one",
    )
    # Whitespace trivia around an unparsable token still normalises away.
    assert semantic_equal(
        ARTIFACT_CATALOG, "candidate_latlon", "  not-a-coordinate ", "not-a-coordinate",
    )


def test_candidate_latlon_conflict_workbook_shows_raw_text(tmp_path: Path) -> None:
    base = _candidate_latlon_baseline("55.619898,37.598040")
    local = _candidate_latlon_baseline("55.619898,37.598040; 55.620000,37.599000")
    remote = _candidate_latlon_baseline("55.619898,37.598040; 59.933333,30.316667")

    result = merge(base, local, remote, deletions_are_meaningful=True)
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", result, PROVENANCE)

    sheet = load_workbook(path)[CONFLICTS_SHEET]
    headers = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]

    assert row[headers.index("This computer")] == "55.619898,37.598040; 55.620000,37.599000"
    assert row[headers.index("Google Drive")] == "55.619898,37.598040; 59.933333,30.316667"


def test_conflict_workbook_shows_original_text_not_normalized_values(tmp_path: Path) -> None:
    base, local, remote = _sides(people="Тоня")
    local.records["020"].fields["people"] = "Тоня, Мама"
    remote.records["020"].fields["people"] = "Мама"  # a genuinely different set

    result = merge(base, local, remote)
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", result, PROVENANCE)

    sheet = load_workbook(path)[CONFLICTS_SHEET]
    headers = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]

    assert row[headers.index("This computer")] == "Тоня, Мама"
    assert row[headers.index("Google Drive")] == "Мама"


# -- The conflict workbook -------------------------------------------------


def _conflicting():
    base, local, remote = _sides(place="Михнево", people="Тоня")
    local.records["020"].fields.update(place="Дача в Михнево", people="Тоня, Аня")
    remote.records["020"].fields["place"] = "Дом в Михнево"
    return merge(base, local, remote)


def test_no_workbook_is_needed_for_an_automatic_merge() -> None:
    base, local, remote = _sides(place="Михнево", people="Тоня")
    local.records["020"].fields["place"] = "Дача"
    remote.records["020"].fields["people"] = "Мама"

    result = merge(base, local, remote)

    assert not result.has_conflicts  # the caller writes the merged workbook directly


def test_workbook_has_the_three_sheets(tmp_path: Path) -> None:
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", _conflicting(), PROVENANCE)

    workbook = load_workbook(path)
    assert set(workbook.sheetnames) == {INFO_SHEET, MERGE_SHEET, CONFLICTS_SHEET}


def test_only_conflicting_cells_are_highlighted(tmp_path: Path) -> None:
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", _conflicting(), PROVENANCE)

    sheet = load_workbook(path)[MERGE_SHEET]
    headers = [cell.value for cell in sheet[1]]
    place = sheet.cell(row=2, column=headers.index("place") + 1)
    people = sheet.cell(row=2, column=headers.index("people") + 1)

    assert place.fill.fgColor.rgb.endswith("FFC7CE")
    # "people" merged automatically — it must not be flagged.
    assert not str(people.fill.fgColor.rgb or "").endswith("FFC7CE")
    assert people.value == "Тоня, Аня"


def test_conflict_cells_carry_a_note_with_all_three_values(tmp_path: Path) -> None:
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", _conflicting(), PROVENANCE)

    sheet = load_workbook(path)[MERGE_SHEET]
    headers = [cell.value for cell in sheet[1]]
    note = sheet.cell(row=2, column=headers.index("place") + 1).comment.text

    assert "CONFLICT" in note
    assert "Михнево" in note and "Дача в Михнево" in note and "Дом в Михнево" in note


def test_conflicts_sheet_has_one_row_per_conflict(tmp_path: Path) -> None:
    result = _conflicting()
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", result, PROVENANCE)

    sheet = load_workbook(path)[CONFLICTS_SHEET]

    assert sheet.max_row == len(result.conflicts) + 1
    headers = [cell.value for cell in sheet[1]]
    assert "Resolution Choice" in headers and "Custom Value" in headers


def test_info_sheet_carries_provenance(tmp_path: Path) -> None:
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", _conflicting(), PROVENANCE)

    text = "\n".join(
        str(value)
        for row in load_workbook(path)[INFO_SHEET].iter_rows(values_only=True)
        for value in row
        if value
    )

    assert "Tonya MacBook" in text and "df0fb26" in text
    assert "sha256:remote" in text and "drive-1" in text


def test_conflict_path_is_separate_from_canonical_files(tmp_path: Path) -> None:
    path = conflict_workbook_path(tmp_path, "run-1", "Архив/A", ARTIFACT_REVIEW)

    assert "_conflicts" in path.parts and "run-1" in path.parts
    assert path.name.endswith(".merge.xlsx")
    assert "/" not in path.name


# -- Resolution ------------------------------------------------------------


def _resolved(result, choices: dict[str, tuple[str, str]]) -> ResolutionSheet:
    return ResolutionSheet(
        conflicts=result.conflicts,
        resolutions={
            key: Resolution(
                record_id=key.split("::")[0], field_name=key.split("::")[1],
                choice=choice, custom_value=custom,
            )
            for key, (choice, custom) in choices.items()
        },
        provenance={},
    )


def test_unresolved_conflicts_are_rejected() -> None:
    result = _conflicting()

    outcome = resolve(result, _resolved(result, {}))

    assert outcome.status is ApplyStatus.INCOMPLETE
    assert outcome.unresolved
    assert "no explicit resolution" in outcome.message


def test_custom_without_a_value_is_rejected() -> None:
    result = _conflicting()
    key = result.conflicts[0].key

    outcome = resolve(result, _resolved(result, {key: (RESOLUTION_CUSTOM, "")}))

    assert outcome.status is ApplyStatus.INCOMPLETE


def test_each_resolution_choice_works() -> None:
    for choice, expected in (
        (RESOLUTION_LOCAL, "Дача в Михнево"),
        (RESOLUTION_DRIVE, "Дом в Михнево"),
        (RESOLUTION_BASE, "Михнево"),
    ):
        result = _conflicting()
        key = result.conflicts[0].key

        outcome = resolve(result, _resolved(result, {key: (choice, "")}))

        assert outcome.ok
        assert outcome.records["020"].value("place") == expected


def test_a_custom_value_is_applied() -> None:
    result = _conflicting()
    key = result.conflicts[0].key

    outcome = resolve(result, _resolved(result, {key: (RESOLUTION_CUSTOM, "Михнево, дача")}))

    assert outcome.ok
    assert outcome.records["020"].value("place") == "Михнево, дача"


def test_automatic_merges_survive_resolution() -> None:
    result = _conflicting()
    key = result.conflicts[0].key

    outcome = resolve(result, _resolved(result, {key: (RESOLUTION_LOCAL, "")}))

    # "people" was never in conflict and keeps its merged value.
    assert outcome.records["020"].value("people") == "Тоня, Аня"


def test_a_workbook_round_trips_through_excel(tmp_path: Path) -> None:
    result = _conflicting()
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", result, PROVENANCE)

    # A person opens it, chooses, saves.
    workbook = load_workbook(path)
    sheet = workbook[CONFLICTS_SHEET]
    headers = [cell.value for cell in sheet[1]]
    sheet.cell(row=2, column=headers.index("Resolution Choice") + 1).value = RESOLUTION_DRIVE
    workbook.save(path)

    outcome = resolve_from_workbook(result, path)

    assert outcome.ok
    assert outcome.records["020"].value("place") == "Дом в Михнево"


# -- Stale remote protection -----------------------------------------------


def test_a_remote_change_during_resolution_aborts() -> None:
    result = _conflicting()
    key = result.conflicts[0].key
    sheet = _resolved(result, {key: (RESOLUTION_LOCAL, "")})

    outcome = resolve(
        result, sheet,
        expected_remote_hash="sha256:R1",
        current_remote_hash=lambda: "sha256:R2",
    )

    assert outcome.status is ApplyStatus.REMOTE_CHANGED
    assert "REMOTE CHANGED SINCE CONFLICT WAS CREATED" in outcome.message
    assert outcome.records == {}


def test_an_unchanged_remote_allows_the_merge() -> None:
    result = _conflicting()
    key = result.conflicts[0].key
    sheet = _resolved(result, {key: (RESOLUTION_LOCAL, "")})

    outcome = resolve(
        result, sheet,
        expected_remote_hash="sha256:R1",
        current_remote_hash=lambda: "sha256:R1",
    )

    assert outcome.ok


def test_the_remote_check_runs_before_anything_is_written() -> None:
    result = _conflicting()
    sheet = _resolved(result, {})  # deliberately incomplete

    outcome = resolve(
        result, sheet,
        expected_remote_hash="sha256:R1",
        current_remote_hash=lambda: "sha256:R2",
    )

    # Incompleteness is reported first; either way nothing is written.
    assert outcome.status is ApplyStatus.INCOMPLETE
    assert outcome.records == {}


def test_a_resolved_workbook_is_archived_not_deleted(tmp_path: Path) -> None:
    path = write_conflict_workbook(tmp_path / "m.merge.xlsx", _conflicting(), PROVENANCE)

    archived = archive_merge_workbook(path)

    assert not path.exists()
    assert archived.exists()
    assert archived.name.endswith(".resolved.xlsx")


def test_the_new_baseline_comes_from_the_resolved_content() -> None:
    result = _conflicting()
    key = result.conflicts[0].key
    outcome = resolve(result, _resolved(result, {key: (RESOLUTION_DRIVE, "")}))

    baseline = outcome.as_baseline(ARTIFACT_REVIEW, "review.xlsx")

    assert baseline.records["020"].value("place") == "Дом в Михнево"
    assert baseline.artifact == ARTIFACT_REVIEW


def test_baselines_round_trip_as_json() -> None:
    baseline = _baseline(place="Михнево", people="Тоня")

    restored = SemanticBaseline.from_dict(baseline.as_dict())

    assert restored.records["020"].fields == baseline.records["020"].fields
    assert restored.order == baseline.order
