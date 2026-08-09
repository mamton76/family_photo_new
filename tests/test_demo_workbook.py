"""Generates the real demo merge workbooks for manual Excel inspection.

These are not throwaway fixtures: they are written with the same production
writer real conflicts use, into ``review-output/_conflicts/demo/`` (the same
``_conflicts`` convention :func:`~photoarchive.merge.workbook.conflict_workbook_path`
uses for genuine runs), so a person can open them directly. No canonical
archive workbook is touched — only files under the ``demo`` run id.

Running this test is what (re)generates the files; the assertions below are
the structural validation the deliverable asks for.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from photoarchive.merge.baseline import ARTIFACT_REVIEW, SemanticBaseline, SemanticRecord
from photoarchive.merge.threeway import ConflictKind, merge, merge_first_sync
from photoarchive.merge.workbook import (
    CONFLICTS_SHEET,
    INFO_SHEET,
    MERGE_SHEET,
    ConflictProvenance,
    write_conflict_workbook,
)
from photoarchive.portable.provenance import app_commit, format_timestamp

REPO_ROOT = Path(__file__).resolve().parents[1]
DEMO_DIR = REPO_ROOT / "review-output" / "_conflicts" / "demo"


def _record(record_id: str, filename: str, **fields: str) -> SemanticRecord:
    return SemanticRecord(record_id=record_id, label=filename, sheet="Review", fields=fields)


def _ordinary_demo_result():
    base = SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={
            "020": _record(
                "020", "020.jpg",
                place="Михнево", people="Тоня", event="", status="REVIEW",
            ),
            "021": _record(
                "021", "021.jpg",
                place="Дача", notes="", status="NEW",
            ),
        },
        order=["020", "021"],
    )
    local = SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={
            "020": _record(
                "020", "020.jpg",
                place="Дача в Михнево", people="Тоня, Мама",
                event="Поездка на дачу",  # changed only here: a plain auto-merge
                status="REVIEW",
            ),
            "021": _record(
                "021", "021.jpg",
                place="Дача", notes="подписано на обороте", status="NEW",
            ),
        },
        order=["020", "021"],
    )
    remote = SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={
            "020": _record(
                "020", "020.jpg",
                place="Дом в Михнево", people="Мама; Тоня", event="", status="REVIEW",
            ),
            "021": _record(
                "021", "021.jpg",
                place="Дача", notes="возможно 1980", status="NEW",
            ),
        },
        order=["020", "021"],
    )
    return merge(base, local, remote, deletions_are_meaningful=False)


def _demo_provenance(**overrides) -> ConflictProvenance:
    fields = dict(
        artifact_path="review-output/Ф-ТоняМам-76-83-разное/review.xlsx",
        artifact_kind=ARTIFACT_REVIEW,
        run_id="demo",
        created_at=format_timestamp(),
        machine_label="Tonya MacBook",
        app_commit=app_commit(REPO_ROOT) or "",
        last_sync_at=format_timestamp(),
        last_sync_machine="Home PC",
        last_sync_run="run-41",
        base_hash="sha256:demo-base",
        local_hash="sha256:demo-local",
        remote_hash="sha256:demo-remote",
        drive_file_id="demo-drive-1",
    )
    fields.update(overrides)
    return ConflictProvenance(**fields)


def test_demo_review_workbook_has_the_expected_shape() -> None:
    result = _ordinary_demo_result()

    # Place is a true conflict; People is not (semantic equality); event is a
    # plain, single-sided auto-merge.
    assert result.conflict_for("020", "place") is not None
    assert result.conflict_for("020", "people") is None
    assert result.conflict_for("020", "event") is None
    assert result.records["020"].value("event") == "Поездка на дачу"
    assert result.records["020"].value("people") in {"Тоня, Мама", "Мама; Тоня"}

    # Notes is a true conflict on the second row.
    assert result.conflict_for("021", "notes") is not None
    assert {conflict.field_name for conflict in result.conflicts} == {"place", "notes"}


def test_generates_demo_review_workbook(capsys) -> None:
    path = DEMO_DIR / "demo-review.merge.xlsx"
    result = _ordinary_demo_result()

    written = write_conflict_workbook(path, result, _demo_provenance())

    assert written == path
    assert path.exists()
    print(f"demo-review.merge.xlsx written to: {path}")

    workbook = load_workbook(path)
    try:
        assert set(workbook.sheetnames) == {INFO_SHEET, MERGE_SHEET, CONFLICTS_SHEET}

        # Conflicts sheet: exactly the two true conflicts, with a working
        # dropdown and a Custom Value column.
        conflicts_sheet = workbook[CONFLICTS_SHEET]
        assert conflicts_sheet.max_row == len(result.conflicts) + 1
        headers = [cell.value for cell in conflicts_sheet[1]]
        assert "Resolution Choice" in headers and "Custom Value" in headers
        assert conflicts_sheet.data_validations.dataValidation, "dropdown is missing"
        formula = conflicts_sheet.data_validations.dataValidation[0].formula1
        assert "LOCAL" in formula and "DRIVE" in formula and "CUSTOM" in formula

        # Merge sheet: only the conflicting cells are highlighted; the
        # auto-merged and semantically-equal cells are not.
        merge_sheet = workbook[MERGE_SHEET]
        merge_headers = [cell.value for cell in merge_sheet[1]]
        assert "Reference" in merge_headers

        def fill_of(row: int, column: str) -> str:
            cell = merge_sheet.cell(row=row, column=merge_headers.index(column) + 1)
            return str(cell.fill.fgColor.rgb or "")

        row_020 = 2  # order=["020", "021"]
        row_021 = 3
        assert fill_of(row_020, "place").endswith("FFC7CE")
        assert not fill_of(row_020, "people").endswith("FFC7CE")
        assert not fill_of(row_020, "event").endswith("FFC7CE")
        assert fill_of(row_021, "notes").endswith("FFC7CE")

        # A conflict cell's comment names Base / This computer / Google Drive.
        note = merge_sheet.cell(
            row=row_020, column=merge_headers.index("place") + 1
        ).comment.text
        assert "Михнево" in note and "Дача в Михнево" in note and "Дом в Михнево" in note

        # Enough context to identify the row without exposing internal ids.
        assert merge_sheet.cell(row=row_020, column=merge_headers.index("Reference") + 1).value == "020.jpg"

        # Frozen headers / sane column widths, carried over from the writer.
        assert merge_sheet.freeze_panes == "D2"
        assert conflicts_sheet.freeze_panes == "A2"
        assert merge_sheet.column_dimensions["A"].width and merge_sheet.column_dimensions["A"].width > 0

        # Info sheet carries provenance a person can act on.
        info_text = "\n".join(
            str(value)
            for row in workbook[INFO_SHEET].iter_rows(values_only=True)
            for value in row
            if value
        )
        assert "Tonya MacBook" in info_text
        assert "review-output/Ф-ТоняМам-76-83-разное/review.xlsx" in info_text
    finally:
        workbook.close()


def _first_sync_demo_result():
    local = SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={"020": _record("020", "020.jpg", place="Дача в Михнево", people="Тоня")},
        order=["020"],
    )
    remote = SemanticBaseline(
        artifact=ARTIFACT_REVIEW,
        records={"020": _record("020", "020.jpg", place="Дом в Михнево", people="Тоня, Мама")},
        order=["020"],
    )
    return merge_first_sync(local, remote)


def test_generates_demo_first_sync_workbook() -> None:
    path = DEMO_DIR / "demo-first-sync.merge.xlsx"
    result = _first_sync_demo_result()
    assert result.has_conflicts
    assert result.conflicts[0].kind is ConflictKind.FIRST_SYNC

    provenance = _demo_provenance(
        first_sync=True, last_sync_at="", last_sync_machine="", last_sync_run="",
        base_hash="",
    )
    written = write_conflict_workbook(path, result, provenance)

    assert written == path
    print(f"demo-first-sync.merge.xlsx written to: {path}")

    workbook = load_workbook(path)
    try:
        info_text = "\n".join(
            str(value)
            for row in workbook[INFO_SHEET].iter_rows(values_only=True)
            for value in row
            if value
        )
        assert "FIRST SYNC" in info_text and "NO COMMON BASELINE" in info_text

        conflicts_sheet = workbook[CONFLICTS_SHEET]
        headers = [cell.value for cell in conflicts_sheet[1]]
        base_cell = conflicts_sheet.cell(row=2, column=headers.index("Base") + 1).value
        assert base_cell == "— no common baseline —"

        formula = conflicts_sheet.data_validations.dataValidation[0].formula1
        assert "BASE" not in formula
        assert "LOCAL" in formula and "DRIVE" in formula and "CUSTOM" in formula
    finally:
        workbook.close()
