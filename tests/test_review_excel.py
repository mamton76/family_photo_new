"""review.xlsx schema and generation tests. Local files only, no cloud."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from photoarchive.models import WorkflowStatus
from photoarchive.review.excel import (
    FINAL_COLUMNS,
    SHEET_NAME,
    SUGGESTED_COLUMNS,
    VISIBLE_COLUMNS,
    ReviewWorkbookService,
    WorkbookPreview,
    build_preview,
    column_index,
    identity_key,
    rows_signature,
)
from photoarchive.review.model import ReviewRow, join_values, split_values

EXPECTED_COLUMNS = (
    "Preview",
    "Filename / Reference",
    "Photo ID",
    "Source Description",
    "Section Context",
    "Source Notes",
    "Description",
    "Suggested Date",
    "Date",
    "Suggested Place",
    "Place",
    "Suggested LatLon",
    "LatLon",
    "Map Link",
    "Suggested People",
    "People",
    "Suggested Tags",
    "Tags",
    "Event",
    "Albums",
    "Status",
    "Review Reason",
    "Notes",
)


def _png(path: Path, size=(60, 40)) -> Path:
    from PIL import Image

    path.parent.mkdir(parents=True, exist_ok=True)
    Image.new("RGB", size, (200, 120, 60)).save(path, format="PNG")
    return path


# -- Schema ---------------------------------------------------------------


def test_visible_columns_match_the_specification() -> None:
    assert VISIBLE_COLUMNS == EXPECTED_COLUMNS


def test_no_precision_columns_exist() -> None:
    assert not any("Precision" in column for column in VISIBLE_COLUMNS)


def test_each_suggested_column_precedes_its_final_column() -> None:
    for suggested in SUGGESTED_COLUMNS:
        final = suggested.replace("Suggested ", "")
        assert column_index(suggested) < column_index(final)


def test_suggested_and_final_columns_are_disjoint() -> None:
    assert not SUGGESTED_COLUMNS & FINAL_COLUMNS


def test_identity_key_ignores_extension_and_case() -> None:
    assert identity_key("20200512_150442.jpg") == "20200512_150442"
    assert identity_key("20200512_150442") == "20200512_150442"
    assert identity_key("IMG_001.JPEG") == "img_001"


def test_list_values_round_trip() -> None:
    assert join_values(["Тоня", "Аня"]) == "Тоня; Аня"
    assert split_values("Тоня; Аня") == ["Тоня", "Аня"]
    assert split_values(None) == []


# -- Generation -----------------------------------------------------------


def _row(reference: str, **kwargs) -> ReviewRow:
    return ReviewRow(reference=reference, **kwargs)


def test_workbook_is_written_with_headers_and_rows(tmp_path: Path) -> None:
    path = tmp_path / "review.xlsx"
    rows = [_row("020", filename="020.jpg", source_description="Тоня на даче")]

    ReviewWorkbookService().write(path, rows)

    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    header = [cell.value for cell in worksheet[1]]
    assert tuple(header) == EXPECTED_COLUMNS
    assert worksheet.cell(row=2, column=column_index("Filename / Reference")).value == "020.jpg"
    assert worksheet.freeze_panes == "C2"
    assert worksheet.auto_filter.ref is not None


def test_cyrillic_survives_a_round_trip(tmp_path: Path) -> None:
    path = tmp_path / "review.xlsx"
    rows = [_row("020", source_description="Тоня Мамаева (3г). Дома.", section_context="Далее Аня")]

    service = ReviewWorkbookService()
    service.write(path, rows)
    reloaded = service.read(path)

    assert reloaded["020"].source_description == "Тоня Мамаева (3г). Дома."
    assert reloaded["020"].section_context == "Далее Аня"


def test_preview_is_embedded_for_present_photos(tmp_path: Path) -> None:
    path = tmp_path / "review.xlsx"
    image = _png(tmp_path / "thumb.png")
    previews = {"020": WorkbookPreview("020", image, 60, 40)}

    ReviewWorkbookService().write(path, [_row("020", filename="020.jpg")], previews)

    workbook = load_workbook(path)
    assert len(workbook[SHEET_NAME]._images) == 1


def test_described_absent_row_has_no_preview(tmp_path: Path) -> None:
    path = tmp_path / "review.xlsx"
    rows = [
        _row("020", filename="020.jpg"),
        _row("021", status=WorkflowStatus.DESCRIBED_ABSENT),
    ]
    previews = {"020": WorkbookPreview("020", _png(tmp_path / "t.png"), 60, 40)}

    ReviewWorkbookService().write(path, rows, previews)

    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    # Only the present photo carries an image; the absent row is still a row.
    assert len(worksheet._images) == 1
    assert worksheet.max_row == 3
    assert worksheet.cell(row=3, column=column_index("Status")).value == "DESCRIBED_ABSENT"


def test_row_order_is_stable_across_rewrites(tmp_path: Path) -> None:
    path = tmp_path / "review.xlsx"
    rows = [_row("020"), _row("021"), _row("022")]
    service = ReviewWorkbookService()

    service.write(path, rows)
    first = [cell.value for cell in load_workbook(path)[SHEET_NAME]["B"]]
    service.write(path, rows)
    second = [cell.value for cell in load_workbook(path)[SHEET_NAME]["B"]]

    assert first == second == [None, "020", "021", "022"][: len(first)] or first == second


def test_read_of_a_missing_workbook_is_empty(tmp_path: Path) -> None:
    assert ReviewWorkbookService().read(tmp_path / "absent.xlsx") == {}


def test_build_preview_downscales_without_touching_the_source(tmp_path: Path) -> None:
    source = _png(tmp_path / "big.png", size=(900, 600))
    before = source.read_bytes()

    size = build_preview(source, tmp_path / "small.png", width_px=180)

    assert size == (180, 120)
    assert (tmp_path / "small.png").exists()
    assert source.read_bytes() == before


def test_build_preview_returns_none_for_unreadable_images(tmp_path: Path) -> None:
    broken = tmp_path / "broken.jpg"
    broken.write_bytes(b"not an image")

    assert build_preview(broken, tmp_path / "out.png") is None


# -- Filling a row in: what the sheet offers the reviewer ------------------


def test_status_column_offers_the_statuses_as_a_list(tmp_path: Path) -> None:
    """An unrecognised status reads back as NEW, so typing must be guarded."""
    path = tmp_path / "review.xlsx"
    ReviewWorkbookService().write(path, [ReviewRow(reference="020", filename="020.jpg")])

    workbook = load_workbook(path)
    try:
        sheet = workbook[SHEET_NAME]
        validations = sheet.data_validations.dataValidation
        assert len(validations) == 1
        validation = validations[0]
        assert "APPROVED" in validation.formula1
        assert "SKIP" in validation.formula1
        # It covers the Status column's data rows, not the header.
        assert str(validation.sqref).startswith(
            get_column_letter(column_index("Status")) + "2"
        )
    finally:
        workbook.close()


def test_source_text_is_attached_to_the_cells_a_reviewer_fills(tmp_path: Path) -> None:
    row = ReviewRow(reference="020", filename="020.jpg")
    row.source_description = "Тоня и Аня у школы"
    row.section_context = "Далее 1990 год"
    row.source_notes = "нет фото"

    path = tmp_path / "review.xlsx"
    ReviewWorkbookService().write(path, [row])

    workbook = load_workbook(path)
    try:
        sheet = workbook[SHEET_NAME]
        comment = sheet.cell(row=2, column=column_index("People")).comment
        assert comment is not None
        assert "Тоня и Аня у школы" in comment.text
        assert "Далее 1990 год" in comment.text
        assert "нет фото" in comment.text
        # Not on every cell: the source columns already show this text.
        assert sheet.cell(row=2, column=column_index("Source Description")).comment is None
    finally:
        workbook.close()


def test_a_row_with_no_source_text_gets_no_comment(tmp_path: Path) -> None:
    path = tmp_path / "review.xlsx"
    ReviewWorkbookService().write(path, [ReviewRow(reference="021", filename="021.jpg")])

    workbook = load_workbook(path)
    try:
        sheet = workbook[SHEET_NAME]
        assert sheet.cell(row=2, column=column_index("People")).comment is None
    finally:
        workbook.close()


def test_a_layout_change_rewrites_an_otherwise_unchanged_workbook() -> None:
    """Otherwise a folder nobody edited would keep the old sheet forever."""
    from photoarchive.review import excel as excel_module

    rows = [ReviewRow(reference="020", filename="020.jpg")]
    before = rows_signature(rows)

    original = excel_module.LAYOUT_VERSION
    try:
        excel_module.LAYOUT_VERSION = original + 1
        assert rows_signature(rows) != before
    finally:
        excel_module.LAYOUT_VERSION = original

    assert rows_signature(rows) == before


def test_the_photo_id_column_round_trips(tmp_path: Path) -> None:
    """The workbook carries the id so one copy can be pointed at another."""
    row = ReviewRow(reference="020", filename="020.jpg")
    row.photo_id = "photo-abc123def456"

    path = tmp_path / "review.xlsx"
    ReviewWorkbookService().write(path, [row])
    read_back = ReviewWorkbookService().read(path)

    assert read_back["020"].photo_id == "photo-abc123def456"


def test_the_fingerprint_survives_rescanning_the_same_picture(tmp_path: Path) -> None:
    """Resolution and brightness move it barely; that is the whole point."""
    from PIL import Image, ImageEnhance

    from photoarchive.review.excel import fingerprint_distance, image_fingerprint

    original = tmp_path / "photo.jpg"
    picture = Image.new("RGB", (400, 300))
    for x in range(400):
        for y in range(300):
            picture.putpixel((x, y), (x % 256, (x * y) % 256, y % 256))
    picture.save(original)

    rescanned = tmp_path / "rescan.jpg"
    smaller = picture.resize((240, 180))
    ImageEnhance.Brightness(smaller).enhance(1.2).save(rescanned)

    assert fingerprint_distance(
        image_fingerprint(original), image_fingerprint(rescanned)
    ) <= 4


def test_an_unreadable_image_has_no_fingerprint(tmp_path: Path) -> None:
    from photoarchive.review.excel import image_fingerprint

    broken = tmp_path / "broken.jpg"
    broken.write_bytes(b"not an image")

    assert image_fingerprint(broken) is None
