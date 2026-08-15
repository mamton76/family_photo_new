"""The iterative loop: learn from review rows, curate catalog.xlsx, rescan."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from photoarchive.catalog.discovery import discover_review_workbooks, read_review_workbooks
from photoarchive.catalog.importer import import_catalog
from photoarchive.catalog.learning import learn_from_rows
from photoarchive.catalog.models import ConfidenceStatus, EntityType
from photoarchive.catalog.service import CatalogService
from photoarchive.catalog.store import DictionaryStore
from photoarchive.geo import LatLon
from photoarchive.models import RemoteSourceItem, WorkflowStatus
from photoarchive.parsing.descriptions import (
    DescriptionEntry,
    ReconciledEntry,
    Reconciliation,
)
from photoarchive.parsing.suggestions import Suggestion, suggest
from photoarchive.review.builder import build_rows
from photoarchive.review.excel import ReviewWorkbookService
from photoarchive.review.model import ReviewRow

MOSCOW = LatLon(55.751244, 37.618423)
OTHER = LatLon(59.934280, 30.335099)


def _store(tmp_path: Path) -> DictionaryStore:
    store = DictionaryStore(tmp_path / "dict.sqlite")
    store.initialize()
    return store


def _row(status=WorkflowStatus.REVIEW, **kwargs) -> ReviewRow:
    row = ReviewRow(reference=kwargs.pop("reference", "020"), **kwargs)
    row.status = status
    return row


# -- Learning source: APPROVED is no longer the gate -----------------------


def test_review_row_with_people_contributes(tmp_path: Path) -> None:
    outcome = learn_from_rows(_store(tmp_path), [_row(people="Тоня Мамаева")])

    assert outcome.new_people == ["Тоня Мамаева"]
    assert outcome.contributing_rows == ["020"]


def test_review_row_with_place_and_latlon_contributes(tmp_path: Path) -> None:
    store = _store(tmp_path)

    outcome = learn_from_rows(
        store, [_row(place="Днепропетровская, Москва", latlon=MOSCOW.format())]
    )

    assert outcome.new_places == ["Днепропетровская, Москва"]
    assert outcome.latlon_added == ["Днепропетровская, Москва"]
    assert store.load().places[0].latlon.format() == MOSCOW.format()


def test_review_row_with_tags_contributes(tmp_path: Path) -> None:
    outcome = learn_from_rows(_store(tmp_path), [_row(tags="Школа")])

    assert outcome.new_tags == ["Школа"]


def test_approved_rows_still_contribute(tmp_path: Path) -> None:
    outcome = learn_from_rows(
        _store(tmp_path), [_row(WorkflowStatus.APPROVED, people="Аня Архангельская")]
    )

    assert outcome.new_people == ["Аня Архангельская"]


def test_error_and_skip_rows_never_teach(tmp_path: Path) -> None:
    rows = [
        _row(WorkflowStatus.ERROR, people="Ошибка"),
        _row(WorkflowStatus.SKIP, people="Пропуск"),
    ]

    outcome = learn_from_rows(_store(tmp_path), rows)

    assert outcome.new_people == []
    assert len(outcome.skipped_rows) == 2


def test_suggested_values_alone_never_teach(tmp_path: Path) -> None:
    store = _store(tmp_path)
    row = _row(
        suggested_people="Машинное Предположение",
        suggested_place="Машинное Место",
        suggested_tags="машина",
    )

    outcome = learn_from_rows(store, [row])

    assert outcome.new_people == []
    assert outcome.new_places == []
    assert store.load().is_empty


def test_comma_separated_people_become_separate_entities(tmp_path: Path) -> None:
    # Reviewers type commas far more often than semicolons.
    store = _store(tmp_path)

    learn_from_rows(store, [_row(people="Тоня Мамаева, Настя Платова")])

    assert {person.canonical_name for person in store.load().people} == {
        "Тоня Мамаева",
        "Настя Платова",
    }


def test_place_commas_are_not_split(tmp_path: Path) -> None:
    store = _store(tmp_path)

    learn_from_rows(store, [_row(place="Днепропетровская, Москва")])

    assert [place.canonical_place for place in store.load().places] == [
        "Днепропетровская, Москва"
    ]


# -- Idempotency ----------------------------------------------------------


def _counts(store: DictionaryStore) -> tuple[int, int, int, int]:
    dictionary = store.load()
    evidence = sum(
        store.evidence_count(EntityType.PERSON, person.canonical_name)
        for person in dictionary.people
    )
    return (
        len(dictionary.people),
        len(dictionary.places),
        len(dictionary.tags),
        evidence,
    )


def test_repeated_unchanged_learn_is_idempotent(tmp_path: Path) -> None:
    store = _store(tmp_path)
    rows = [_row(people="Тоня Мамаева", place="Михнево", tags="дача", source_description="Тоня")]

    learn_from_rows(store, rows)
    first = _counts(store)
    learn_from_rows(store, rows)
    learn_from_rows(store, rows)

    assert _counts(store) == first


def test_later_edits_add_only_new_knowledge(tmp_path: Path) -> None:
    store = _store(tmp_path)
    learn_from_rows(store, [_row(people="Тоня Мамаева")])

    outcome = learn_from_rows(
        store, [_row(people="Тоня Мамаева, Настя Платова")]
    )

    assert outcome.new_people == ["Настя Платова"]
    assert len(store.load().people) == 2


# -- Blank-final propagation ----------------------------------------------


def _reconciliation(reference="020", text="описание"):
    entry = DescriptionEntry(reference=reference, paragraphs=(text,), text=text)
    photo = RemoteSourceItem(
        name=f"{reference}.jpg", relative_path=f"{reference}.jpg", is_directory=False
    )
    return Reconciliation(entries=(ReconciledEntry(entry=entry, photo=photo),))


FULL = Suggestion(
    date="1979", place="Михнево", latlon=MOSCOW.format(),
    people=("Тоня Мамаева",), tags=("дача",),
)


def _existing(**kwargs) -> dict[str, ReviewRow]:
    row = ReviewRow(reference="020", **kwargs)
    row.status = WorkflowStatus.REVIEW
    return {"020": row}


def test_blank_final_fields_are_filled_from_suggestions() -> None:
    outcome, _ = build_rows(
        _reconciliation(), {"020": FULL}, existing=_existing(), states={}
    )
    row = outcome.rows[0]

    assert row.date == "1979"
    assert row.place == "Михнево"
    assert row.latlon == MOSCOW.format()
    assert row.people == "Тоня Мамаева"
    assert row.tags == "дача"
    # The five suggested fields, plus the caption seeded from the source text.
    assert len(outcome.autofilled) == 6
    assert row.description == "описание"


def test_non_empty_final_values_are_never_overwritten() -> None:
    existing = _existing(
        date="1981", place="Дача", latlon=OTHER.format(),
        people="Кто-то Другой", tags="свой-тег",
    )

    outcome, _ = build_rows(_reconciliation(), {"020": FULL}, existing=existing, states={})
    row = outcome.rows[0]

    assert row.date == "1981"
    assert row.place == "Дача"
    assert row.latlon == OTHER.format()
    assert row.people == "Кто-то Другой"
    assert row.tags == "свой-тег"
    # Only the untouched caption is seeded; every typed value stands.
    assert outcome.autofilled == ["020.description"]
    assert outcome.preserved == ["020"]


def test_mixed_row_fills_only_the_blanks() -> None:
    # The exact scenario from the specification: Place typed, LatLon blank.
    existing = _existing(place="Днепропетровская, Москва")

    outcome, _ = build_rows(_reconciliation(), {"020": FULL}, existing=existing, states={})
    row = outcome.rows[0]

    assert row.place == "Днепропетровская, Москва"
    assert row.latlon == MOSCOW.format()
    assert "020.latlon" in outcome.autofilled
    assert "020.place" not in outcome.autofilled


def test_blank_finals_stay_blank_without_suggestions() -> None:
    outcome, _ = build_rows(
        _reconciliation(), {"020": Suggestion()}, existing=_existing(), states={}
    )

    assert outcome.rows[0].place == ""
    # No suggestions, so nothing but the caption is written.
    assert outcome.autofilled == ["020.description"]


# -- Full propagation loop ------------------------------------------------


def test_learn_then_rescan_propagates_coordinates(tmp_path: Path) -> None:
    store = _store(tmp_path)
    # One reviewed row teaches the place and its coordinates.
    learn_from_rows(store, [_row(place="Михнево", latlon=MOSCOW.format())])

    # A different photo whose description mentions the place.
    suggestion = suggest("1979 май. На даче в Михнево.", store.load())
    assert suggestion.place == "Михнево"
    assert suggestion.latlon == MOSCOW.format()

    outcome, _ = build_rows(
        _reconciliation("021"), {"021": suggestion},
        existing={"021": ReviewRow(reference="021")}, states={},
    )
    row = outcome.rows[0]

    assert row.place == "Михнево"
    assert row.latlon == MOSCOW.format()


def test_propagation_preserves_a_typed_place(tmp_path: Path) -> None:
    store = _store(tmp_path)
    learn_from_rows(store, [_row(place="Михнево", latlon=MOSCOW.format())])
    suggestion = suggest("На даче в Михнево.", store.load())

    existing = {"021": ReviewRow(reference="021", place="Моя дача")}
    outcome, _ = build_rows(_reconciliation("021"), {"021": suggestion}, existing=existing, states={})

    assert outcome.rows[0].place == "Моя дача"
    assert outcome.rows[0].latlon == MOSCOW.format()


# -- Multiple-place ambiguity ---------------------------------------------


def test_two_matching_places_leave_the_suggestion_empty(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)
    store.add_place("Валаам", OTHER)

    suggestion = suggest("Сначала Михнево, потом Валаам.", store.load())

    assert suggestion.place == ""
    assert suggestion.latlon == ""
    assert set(suggestion.ambiguous_places) == {"Михнево", "Валаам"}


def test_a_single_match_is_not_ambiguous(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)

    suggestion = suggest("На даче в Михнево.", store.load())

    assert suggestion.place == "Михнево"
    assert suggestion.ambiguous_places == ()


# -- Bidirectional catalog ------------------------------------------------


def _export_and_edit(store: DictionaryStore, tmp_path: Path, edits) -> Path:
    path, _ = CatalogService().export(store, tmp_path)
    workbook = load_workbook(path)
    edits(workbook)
    workbook.save(path)
    return path


def _cell(sheet, row: int, column: str):
    headers = [cell.value for cell in sheet[1]]
    return sheet.cell(row=row, column=headers.index(column) + 1)


def test_canonical_rename_updates_the_same_entity(tmp_path: Path) -> None:
    store = _store(tmp_path)
    learn_from_rows(store, [_row(people="Тоня Мамаева")])
    original_id = store.load().people[0].person_id

    path = _export_and_edit(
        store, tmp_path,
        lambda wb: setattr(_cell(wb["People"], 2, "canonical_name"), "value", "Антонина Мамаева"),
    )
    outcome = import_catalog(store, path)

    people = store.load().people
    assert len(people) == 1
    assert people[0].canonical_name == "Антонина Мамаева"
    assert people[0].person_id == original_id
    assert outcome.entities_renamed == ["Тоня Мамаева -> Антонина Мамаева"]


def test_manually_added_confirmed_alias_imports(tmp_path: Path) -> None:
    store = _store(tmp_path)
    learn_from_rows(store, [_row(people="Антонина Мамаева")])

    path = _export_and_edit(
        store, tmp_path,
        lambda wb: setattr(_cell(wb["People"], 2, "confirmed_aliases"), "value", "Тоня"),
    )
    outcome = import_catalog(store, path)

    assert "Тоня" in store.load().people[0].confirmed_aliases
    assert outcome.aliases_confirmed == ["Тоня -> Антонина Мамаева"]


def test_candidate_alias_promotion_imports(tmp_path: Path) -> None:
    store = _store(tmp_path)
    person = store.add_person("Антонина Мамаева")
    store.add_alias(EntityType.PERSON, person, "Тонечка", ConfidenceStatus.CANDIDATE)

    def promote(workbook):
        sheet = workbook["People"]
        _cell(sheet, 2, "confirmed_aliases").value = "Тонечка"
        _cell(sheet, 2, "candidate_aliases").value = ""

    path = _export_and_edit(store, tmp_path, promote)
    outcome = import_catalog(store, path)

    reloaded = store.load().people[0]
    assert reloaded.confirmed_aliases == ("Тонечка",)
    assert reloaded.candidate_aliases == ()
    assert outcome.promotions == ["Тонечка -> Антонина Мамаева"]


def test_evidence_survives_a_promotion_via_catalog(tmp_path: Path) -> None:
    store = _store(tmp_path)
    learn_from_rows(
        store,
        [_row(people="Антонина Мамаева", source_description="Тоня Мамаева. Сережа Мамаев.")],
    )
    before = store.evidence_count(EntityType.PERSON, "Антонина Мамаева")

    path = _export_and_edit(
        store, tmp_path,
        lambda wb: setattr(_cell(wb["People"], 2, "confirmed_aliases"), "value", "Тоня Мамаева"),
    )
    import_catalog(store, path)

    assert store.evidence_count(EntityType.PERSON, "Антонина Мамаева") >= before


def test_manual_place_latlon_edit_imports(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево")

    path = _export_and_edit(
        store, tmp_path,
        lambda wb: setattr(_cell(wb["Places"], 2, "latlon"), "value", MOSCOW.format()),
    )
    outcome = import_catalog(store, path)

    assert store.load().places[0].latlon.format() == MOSCOW.format()
    assert len(outcome.latlon_updated) == 1


def test_candidate_latlon_promotion_imports(tmp_path: Path) -> None:
    store = _store(tmp_path)
    place = store.add_place("Михнево", MOSCOW)
    store.propose_place_latlon(place, OTHER)

    path = _export_and_edit(
        store, tmp_path,
        lambda wb: setattr(_cell(wb["Places"], 2, "latlon"), "value", OTHER.format()),
    )
    outcome = import_catalog(store, path)

    reloaded = store.load().places[0]
    assert reloaded.latlon.format() == OTHER.format()
    assert reloaded.candidate_latlon == []
    assert len(outcome.latlon_promoted) == 1


def test_invalid_latlon_does_not_corrupt_confirmed_data(tmp_path: Path) -> None:
    store = _store(tmp_path)
    store.add_place("Михнево", MOSCOW)

    path = _export_and_edit(
        store, tmp_path,
        lambda wb: setattr(_cell(wb["Places"], 2, "latlon"), "value", "где-то рядом"),
    )
    outcome = import_catalog(store, path)

    assert store.load().places[0].latlon.format() == MOSCOW.format()
    assert len(outcome.invalid) == 1


def test_second_unchanged_import_is_idempotent(tmp_path: Path) -> None:
    store = _store(tmp_path)
    learn_from_rows(store, [_row(people="Антонина Мамаева", place="Михнево", tags="дача")])
    path, _ = CatalogService().export(store, tmp_path)

    first = import_catalog(store, path)
    second = import_catalog(store, path)

    assert second.changed == 0
    assert first.changed >= 0
    assert len(store.load().people) == 1


def test_importing_a_missing_catalog_is_harmless(tmp_path: Path) -> None:
    outcome = import_catalog(_store(tmp_path), tmp_path / "nope.xlsx")

    assert outcome.changed == 0


# -- Workbook discovery ---------------------------------------------------


def test_discovery_finds_workbooks_recursively(tmp_path: Path) -> None:
    service = ReviewWorkbookService()
    for folder in ("Root A", "Root B/1988"):
        service.write(tmp_path / folder / "review.xlsx", [ReviewRow(reference="020")])

    found = discover_review_workbooks(tmp_path)

    assert len(found) == 2


def test_discovery_skips_excel_lock_files(tmp_path: Path) -> None:
    ReviewWorkbookService().write(tmp_path / "a" / "review.xlsx", [ReviewRow(reference="1")])
    (tmp_path / "a" / "~$review.xlsx").write_bytes(b"lock")

    assert len(discover_review_workbooks(tmp_path)) == 1


def test_reading_workbooks_returns_rows(tmp_path: Path) -> None:
    row = ReviewRow(reference="020", people="Тоня Мамаева")
    ReviewWorkbookService().write(tmp_path / "f" / "review.xlsx", [row])

    workbooks = read_review_workbooks(tmp_path)

    assert len(workbooks) == 1
    assert workbooks[0].rows[0].people == "Тоня Мамаева"


def test_discovery_of_a_missing_directory_is_empty(tmp_path: Path) -> None:
    assert discover_review_workbooks(tmp_path / "absent") == []


# -- Row identity scope ---------------------------------------------------


def test_same_stem_in_different_folders_does_not_collide() -> None:
    from photoarchive.review.excel import scoped_row_key

    first = scoped_row_key("root-1", "folder A", "001.jpg")
    second = scoped_row_key("root-1", "folder B", "001.jpg")

    assert first != second


def test_same_stem_in_different_source_roots_does_not_collide() -> None:
    from photoarchive.review.excel import scoped_row_key

    assert scoped_row_key("root-1", "", "001.jpg") != scoped_row_key("root-2", "", "001.jpg")


def test_reference_and_filename_share_identity_within_a_folder() -> None:
    from photoarchive.review.excel import scoped_row_key

    # DESCRIBED_ABSENT "20200512_150442" and the photo that later appears.
    absent = scoped_row_key("root-1", "1988", "20200512_150442")
    present = scoped_row_key("root-1", "1988", "20200512_150442.jpg")

    assert absent == present


def test_state_rows_are_scoped_per_folder(tmp_path: Path) -> None:
    from photoarchive.review.builder import RowState
    from photoarchive.state import StateRepository

    state = StateRepository(tmp_path / "archive.sqlite")
    state.initialize()

    state.save_row_states("root-1", "folder A", {"001": RowState("001", photo_hash="a")})
    state.save_row_states("root-1", "folder B", {"001": RowState("001", photo_hash="b")})

    assert state.load_row_states("root-1", "folder A")["001"].photo_hash == "a"
    assert state.load_row_states("root-1", "folder B")["001"].photo_hash == "b"


# -- Merging duplicate entities via catalog.xlsx ---------------------------


def test_alias_naming_another_entity_merges_them(tmp_path: Path) -> None:
    store = _store(tmp_path)
    learn_from_rows(store, [_row(place="Дома на Днепропетровской")])
    learn_from_rows(store, [_row(reference="021", place="Дом на Днепропетровской")])
    assert len(store.load().places) == 2

    def merge(workbook):
        sheet = workbook["Places"]
        headers = [cell.value for cell in sheet[1]]
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row, headers.index("canonical_place") + 1).value
            if name == "Дома на Днепропетровской":
                sheet.cell(row, headers.index("confirmed_aliases") + 1).value = (
                    "Дом на Днепропетровской"
                )
        for row in range(sheet.max_row, 1, -1):
            if sheet.cell(row, headers.index("canonical_place") + 1).value == (
                "Дом на Днепропетровской"
            ):
                sheet.delete_rows(row)

    path = _export_and_edit(store, tmp_path, merge)
    outcome = import_catalog(store, path)

    places = store.load().places
    assert len(places) == 1
    assert places[0].canonical_place == "Дома на Днепропетровской"
    assert "Дом на Днепропетровской" in places[0].confirmed_aliases
    assert outcome.merged == ["Дом на Днепропетровской -> Дома на Днепропетровской"]


def test_a_merged_spelling_is_not_recreated_by_the_next_learn(tmp_path: Path) -> None:
    # The review rows still say "Дом …"; learning must resolve it to the
    # surviving place instead of recreating the duplicate.
    store = _store(tmp_path)
    place_id = store.add_place("Дома на Днепропетровской")
    store.add_alias(
        EntityType.PLACE, place_id, "Дом на Днепропетровской", ConfidenceStatus.CONFIRMED
    )

    learn_from_rows(store, [_row(place="Дом на Днепропетровской")])

    places = store.load().places
    assert len(places) == 1
    assert places[0].canonical_place == "Дома на Днепропетровской"


def test_merging_keeps_the_survivor_coordinates(tmp_path: Path) -> None:
    store = _store(tmp_path)
    keep = store.add_place("Дома на Днепропетровской", MOSCOW)
    absorb = store.add_place("Дом на Днепропетровской")

    store.merge_entities(EntityType.PLACE, keep, absorb)

    places = store.load().places
    assert len(places) == 1
    assert places[0].latlon.format() == MOSCOW.format()


def test_merging_adopts_coordinates_the_survivor_lacked(tmp_path: Path) -> None:
    store = _store(tmp_path)
    keep = store.add_place("Дома на Днепропетровской")
    absorb = store.add_place("Дом на Днепропетровской", MOSCOW)

    store.merge_entities(EntityType.PLACE, keep, absorb)

    assert store.load().places[0].latlon.format() == MOSCOW.format()


def test_merging_keeps_the_absorbed_evidence(tmp_path: Path) -> None:
    store = _store(tmp_path)
    learn_from_rows(store, [_row(place="Дома на Днепропетровской")])
    learn_from_rows(store, [_row(reference="021", place="Дом на Днепропетровской")])
    keep = store.find_by_canonical(EntityType.PLACE, "Дома на Днепропетровской")
    absorb = store.find_by_canonical(EntityType.PLACE, "Дом на Днепропетровской")

    store.merge_entities(EntityType.PLACE, keep, absorb)

    evidence = store.evidence_for(EntityType.PLACE, "Дома на Днепропетровской")
    assert len(evidence) >= 2


def test_repeated_merge_import_is_idempotent(tmp_path: Path) -> None:
    store = _store(tmp_path)
    keep = store.add_place("Дома на Днепропетровской")
    store.add_alias(
        EntityType.PLACE, keep, "Дом на Днепропетровской", ConfidenceStatus.CONFIRMED
    )
    path, _ = CatalogService().export(store, tmp_path)

    import_catalog(store, path)
    second = import_catalog(store, path)

    assert second.merged == []
    assert len(store.load().places) == 1


def test_merging_an_entity_into_itself_does_nothing(tmp_path: Path) -> None:
    store = _store(tmp_path)
    place = store.add_place("Михнево")

    assert store.merge_entities(EntityType.PLACE, place, place) is None
    assert len(store.load().places) == 1


def test_the_caption_starts_from_the_source_text_but_stays_the_reviewers() -> None:
    """`Description` is seeded, then owned: a rewrite is never undone."""
    outcome, states = build_rows(
        _reconciliation(), {"020": Suggestion()}, existing=_existing(), states={}
    )
    row = outcome.rows[0]
    assert row.description == "описание"

    row.description = "Тоня у дома на Днепропетровской, лето"
    rescan, _ = build_rows(
        _reconciliation(),
        {"020": Suggestion()},
        existing={"020": row},
        states=states,
    )

    assert rescan.rows[0].description == "Тоня у дома на Днепропетровской, лето"
    assert "020.description" not in rescan.autofilled
    # The machine's own copy of the source text is refreshed as always.
    assert rescan.rows[0].source_description == "описание"
